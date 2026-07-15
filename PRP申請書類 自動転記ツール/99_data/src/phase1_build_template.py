# -*- coding: utf-8 -*-
"""
Phase1（PRP・複数シート版）: マッピング定義(mapping.json)を生成する。

役割:
  1) テンプレ(2種関節系PRP.xlsx / 3種筋腱靭帯系PRP.xlsx)の全シートの「緑フォント」セルを走査。
  2) 下記のキュレーション済みマッピング(MAPPING)と突き合わせ、各緑セルに転記元(source)を割当。
     - 割当済み   : ヒアリング項目 or クロスシート参照(略歴書/PRPメーカー/審査委員会)
     - 未割当(緑) : source = {"t":"unknown"} として残し、要マッピングで可視化
  3) 99_data/mapping.json として書き出す（transcribe.py が読む外部定義）。

※「{変数名}への置換」は本ツールでは行わず、緑セルへ実値を直接書き込む方式。
   置換の正本＝この mapping.json（要件2の『正確なマッピング定義』に相当）。

ソース種別(source.t):
  hearing  : ヒアリングシート（PRP）を ラベル+セクション+出現順 で検索
  sheet    : ヒアリングブックの別シートのセルを直接参照（例 略歴書（PRP）!B3）
  committee: 委員会名(ヒアリング)→ 審査委員会シートで認定番号を取得
  kit      : PRPキット名(ヒアリング)→ PRPメーカーシートの採取方法(C)/加工方法(D)
  date     : 実行日を和暦で挿入
  fixed    : 固定文字列
  unknown  : 未割当（要手動マッピング。緑セルだが転記元が未定義）
変換(tf): text(既定)/ postal(〒付与)/ prefix(接頭辞付与)
"""
import openpyxl, os, json, io

# スクリプトの場所: <ルート>/99_data/src/phase1_build_template.py
SRC_DIR = os.path.dirname(os.path.abspath(__file__))   # 99_data/src
DATA = os.path.dirname(SRC_DIR)                         # 99_data
BASE = os.path.dirname(DATA)                            # ツールのルート
TPL = os.path.join(DATA, "テンプレート")
MAP_DIR = os.path.join(DATA, "マッピング")
HEARING_SHEET = "ヒアリングシート（PRP）"

# ===== 文書（テンプレ）定義 =====
DOCS = {
    "2種関節系PRP": dict(
        template="2種関節系PRP.xlsx",
        sheet_plan="01.【様式第一の二】 再生医療等提供計画 (治療) 整形外科",
        sheet_rireki="3.医師略歴書",
        sheet_shorui2="2 提供する再生医療等の詳細を記した書類（2種）",
        sheet_setsumei="4.5 再生医療等を受けられる患者様に対する説明書（2種）",
        sheet_gaiyo="8 特定細胞加工物概要書（2種）2026",
        sheet_hyojun="9 特定細胞加工物標準書（2種）2026",
        sheet_assent="アセント文書（2種_関節",
        sheet_doui="再生医療等提供計画の情報の公表に関する同意書2026",
        meiyaku="再生医療①（右記タブから選択）",   # 名称の入力元
        plan_cells="2",
    ),
    "3種筋腱靭帯系PRP": dict(
        template="3種筋腱靭帯系PRP.xlsx",
        sheet_plan="01.【様式第一の二】 再生医療等提供計画 (治療) 整形外科",
        sheet_rireki="3.医師略歴書(迫 浩輔）",
        sheet_shorui2="2 提供する再生医療等の詳細を記した書類（3種）",
        sheet_setsumei="4.5 再生医療等を受けられる患者様に対する説明書（3種）",
        sheet_gaiyo="8 特定細胞加工物概要書（3種）",
        sheet_hyojun="9 特定細胞加工物標準書（3種）2026",
        sheet_assent="アセント文書（3種_腱",
        sheet_doui="再生医療等提供計画の情報の公表に関する同意書2026",
        meiyaku="再生医療②（右記タブから選択）",
        plan_cells="3",
    ),
}

# ===== ソースの短縮ビルダー =====
def H(label, section="", occ=1):
    return {"t": "hearing", "label": label, "section": section, "occ": occ}
def S(sheet, cell):
    return {"t": "sheet", "sheet": sheet, "cell": cell}
def KIT(part, idx):  # part: 採取/加工, idx: 1..3
    return {"t": "kit", "part": part, "idx": idx}
COMMITTEE = {"t": "committee"}

# よく使うヒアリング・ソース
SRC = dict(
    med_name=H("医療機関/名称（診療所開設届上）", "法人/医療機関"),
    med_post=H("医療機関/郵便番号（診療所開設届上）", "法人/医療機関"),
    med_addr=H("医療機関/住所（診療所開設届上）", "法人/医療機関"),
    kaisetsu=H("開設者（診療所開設届上）※医療法人の場合", "法人/医療機関"),
    resp_name=H("氏名", "実施責任者"),
    resp_tel=H("電話番号", "実施責任者"),
    resp_mail=H("メールアドレス", "実施責任者"),
    jimu_name=H("氏名", "事務担当者"),
    jimu_tel=H("電話番号", "事務担当者"),
    jimu_mail=H("メールアドレス", "事務担当者"),
    dr1=H("再生医療を行う医師の氏名", "人員", 1),
    dr2=H("再生医療を行う医師の氏名", "人員", 2),
    kyukyu=H("救急医療に必要な施設又は設備の確認", "救急対応"),
    seizo_name=H("特定細胞加工物製造事業者の名称", "細胞加工施設"),
    shisetsu_no=H("施設番号", "細胞加工施設"),
    kako_name=H("細胞培養加工施設の名称", "細胞加工施設"),
    iinkai=H("認定再生医療等委員会の名称", "審査委員会"),
    toiawase=H("患者からの問い合わせ先", "人員"),
)

# ===== SOP（衛生・製造・品質基準書）=====
#   SOP.xlsx は17シート構成（基準書A〜P＋記録表）。クライアント固有は各基準書冒頭 A8 の
#   医療機関名のみ（緑マークは無いためキュレーションで直接指定）。手順書D〜Pは汎用で対象外。
SOP_TEMPLATE = "SOP.xlsx"
SOP_SHEETS = ["10 【A】衛生管理基準書", "11 【B】製造管理基準書", "12 【C】品質管理基準書"]
SOP_FIELDS = [   # (sheet, cell, var, source, tf, note)
    (s, "A8", "SOP_医療機関名", SRC["med_name"], "text", "各SOP冒頭の医療機関名")
    for s in SOP_SHEETS
]


# ===== 様式第一の二（提供計画）シート1のフィールド（2種/3種で出力先セルが異なる）=====
#   (var, cell_2, cell_3, source, tf, note)
PLAN_FIELDS = [
    ("医療機関_名称",   "N11", "N11", SRC["med_name"], "text", ""),
    ("医療機関_住所",   "N14", "N14", SRC["med_addr"], "text", ""),
    ("管理者_氏名",     "N17", "N17", SRC["kaisetsu"], "text", "確認対象:法人=役職+氏名/個人=氏名のみ"),
    ("再生医療等_名称", "L27", "L27", "MEIYAKU",       "text", "2種=再生医療①/3種=再生医療②"),
    ("実施責任者_氏名",   "L44", "L42", SRC["resp_name"], "text", ""),
    ("実施責任者_所属機関", "L45", "L43", SRC["med_name"], "text", ""),
    ("実施責任者_所属部署", "L46", "L44", SRC["med_name"], "text", "確認対象:部署欄なし→医療機関名で代用"),
    ("実施責任者_郵便番号", "L47", "L45", SRC["med_post"], "postal", ""),
    ("実施責任者_住所",   "L48", "L46", SRC["med_addr"], "text", ""),
    ("実施責任者_電話",   "L49", "L47", SRC["resp_tel"], "text", ""),
    ("実施責任者_メール", "L50", "L48", SRC["resp_mail"], "text", ""),
    ("事務担当者_氏名",   "L51", "L49", SRC["jimu_name"], "text", ""),
    ("事務担当者_所属機関", "L52", "L50", SRC["med_name"], "text", ""),
    ("事務担当者_所属部署", "L53", "L51", SRC["med_name"], "text", "確認対象:部署欄なし→医療機関名で代用"),
    ("事務担当者_郵便番号", "L54", "L52", SRC["med_post"], "postal", ""),
    ("事務担当者_住所",   "L55", "L53", SRC["med_addr"], "text", ""),
    ("事務担当者_電話",   "L56", "L54", SRC["jimu_tel"], "text", ""),
    ("事務担当者_FAX",    "L57", "L55", {"t": "unknown"}, "text", "確認対象:ヒアリングにFAX欄なし"),
    ("事務担当者_メール", "L58", "L56", SRC["jimu_mail"], "text", ""),
    ("医師1_氏名",       "L60", "L58", SRC["dr1"], "text", ""),
    ("医師1_所属機関部署", "L61", "L59", SRC["med_name"], "text", ""),
    ("医師1_役職",       "L62", "L60", {"t": "fixed", "value": "院長"}, "text", "確認対象:既定値院長"),
    ("医師2_氏名",       None,  "L62", SRC["dr2"], "text", "確認対象:2人目。空なら未転記"),
    ("医師2_所属機関部署", None,  "L63", SRC["med_name"], "text", "確認対象:医師2の所属"),
    ("救急_設備内容",     "L68", "L66", SRC["kyukyu"], "text", ""),
    ("加工方法",         "L92", "L92", KIT("加工", 0), "text", "確認対象:PRPキット→加工方法を連結"),
    ("製造事業者_名称",   "L97", "L97", SRC["seizo_name"], "text", ""),
    ("加工施設_施設番号", "L98", "L98", SRC["shisetsu_no"], "text", ""),
    ("加工施設_名称",     "L99", "L99", SRC["kako_name"], "text", ""),
    ("委員会_名称",       "L155", "L152", SRC["iinkai"], "text", ""),
    ("委員会_認定番号",   "L156", "L153", COMMITTEE, "text", "委員会名→審査委員会シートで番号取得"),
]

# ===== 医師略歴書（ヒアリング 略歴書（PRP）シートからのクロスシート参照）=====
RIREKI = "略歴書（PRP）"
RIREKI_FIELDS = [   # (cell, var, source, tf, note)
    ("G4",  "医師_氏名カナ",   S(RIREKI, "B2"), "text", ""),
    ("G5",  "医師_氏名",       S(RIREKI, "B3"), "text", ""),
    ("G6",  "医師_生年月日",   S(RIREKI, "B4"), "text", ""),
    ("G7",  "医師_所属",       S(RIREKI, "B5"), "text", ""),
    ("G8",  "医師_役職",       S(RIREKI, "B6"), "text", ""),
    ("G9",  "医師_学歴",       {"t": "unknown"}, "text", "確認対象:ヒアリング略歴書に学歴欄なし"),
    ("G10", "医師_医師免許",   S(RIREKI, "B7"), "text", "確認対象:略歴書B7(医師免許+職歴)"),
    ("G11", "医師_認定資格",   S(RIREKI, "B10"), "text", ""),
    ("G12", "医師_職歴",       S(RIREKI, "B7"), "text", "確認対象:略歴書B7(医師免許+職歴)"),
    ("G13", "医師_専門分野",   S(RIREKI, "B8"), "text", ""),
    ("G14", "医師_所属学会",   S(RIREKI, "B9"), "text", ""),
    ("G15", "医師_臨床経験",   S(RIREKI, "B11"), "text", ""),
]

# ===== 患者説明書（3種で緑マークあり）=====
SETSUMEI_FIELDS = [   # (cell, var, source, tf, note)
    ("G10", "説明_医療機関名", SRC["med_name"], "text", ""),
    ("G11", "説明_管理者",     SRC["kaisetsu"], "text", ""),
    ("G12", "説明_実施責任者", SRC["resp_name"], "text", ""),
    ("G13", "説明_医師",       SRC["dr1"], "text", ""),
    ("A190", "説明_署名医療機関1", SRC["med_name"], "text", ""),
    ("A214", "説明_署名医療機関2", SRC["med_name"], "text", ""),
    # 委員会・問い合わせ先の埋め込み行は要マッピング(unknown)で可視化
    ("A133", "説明_委員会行",   {"t": "unknown"}, "text", "確認対象:委員会名+番号の埋込行"),
    ("A144", "説明_問合せ連絡先", {"t": "unknown"}, "text", "確認対象:問い合わせ先の埋込行"),
]

# ===== 概要書 / 標準書 / 同意書 / アセント（3種中心）=====
GAIYO_FIELDS = [("D16", "概要_医療機関", SRC["med_name"], "prefix:名称）", "確認対象:名称+所在地の結合セル")]
HYOJUN_FIELDS = [("F17", "標準_医療機関", SRC["med_name"], "text", "")]
DOUI_FIELDS = [
    ("A127", "同意_日付",     {"t": "date"}, "text", "確認対象:同意日(実行日を和暦)"),
    ("A128", "同意_医療機関", SRC["med_name"], "prefix:再生医療等提供機関　名称　", ""),
    ("A129", "同意_住所",     SRC["med_addr"], "prefix:住所　", ""),
    ("A130", "同意_管理者",   SRC["kaisetsu"], "prefix:管理者　氏名　", ""),
]
ASSENT_FIELDS = [("A5", "アセント_医療機関", SRC["med_name"], "text", "確認対象:アセント様式は要目視")]

# ===== 書類2 PRP製造方法（3種で緑マークあり。PRPメーカーからのクロスシート）=====
SHORUI2_FIELDS = [
    ("A34", "書2_採取方法1", KIT("採取", 1), "text", "確認対象:PRPキット①の採取方法"),
    ("A37", "書2_採取方法2", KIT("採取", 2), "text", "確認対象:PRPキット②の採取方法"),
    ("A49", "書2_加工方法1", KIT("加工", 1), "text", "確認対象:PRPキット①の加工方法"),
    ("A57", "書2_加工方法2", KIT("加工", 2), "text", "確認対象:PRPキット②の加工方法"),
]


def build_curated(doc_key):
    """1文書ぶんのキュレーション・マッピング(list of entry)を組み立てる。"""
    d = DOCS[doc_key]
    cellkey = "cell_2" if d["plan_cells"] == "2" else "cell_3"
    entries = []

    def add(sheet, cell, var, src, tf, note):
        if not cell:
            return
        if src == "MEIYAKU":
            src = H(d["meiyaku"], "再生医療の名称")
        entries.append(dict(sheet=sheet, cell=cell, var=var, source=src, tf=tf, note=note))

    # 様式第一の二
    for var, c2, c3, src, tf, note in PLAN_FIELDS:
        add(d["sheet_plan"], c2 if d["plan_cells"] == "2" else c3, var, src, tf, note)
    # 医師略歴書
    for cell, var, src, tf, note in RIREKI_FIELDS:
        add(d["sheet_rireki"], cell, var, src, tf, note)
    # 以下は 3種テンプレで緑マークされている書類（2種は未マークのため緑走査で拾えた分だけ）
    for cell, var, src, tf, note in SETSUMEI_FIELDS:
        add(d["sheet_setsumei"], cell, var, src, tf, note)
    for cell, var, src, tf, note in GAIYO_FIELDS:
        add(d["sheet_gaiyo"], cell, var, src, tf, note)
    for cell, var, src, tf, note in HYOJUN_FIELDS:
        add(d["sheet_hyojun"], cell, var, src, tf, note)
    for cell, var, src, tf, note in DOUI_FIELDS:
        add(d["sheet_doui"], cell, var, src, tf, note)
    for cell, var, src, tf, note in ASSENT_FIELDS:
        add(d["sheet_assent"], cell, var, src, tf, note)
    for cell, var, src, tf, note in SHORUI2_FIELDS:
        add(d["sheet_shorui2"], cell, var, src, tf, note)
    return entries


def is_green(c):
    f = c.font
    return bool(f and f.color and f.color.rgb and isinstance(f.color.rgb, str)
               and f.color.rgb.upper().endswith("00B050"))


def scan_green(template_path):
    """テンプレの全シートの緑セル座標を {sheet: set(cells)} で返す。"""
    wb = openpyxl.load_workbook(template_path)
    found = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cells = set()
        for row in ws.iter_rows():
            for c in row:
                if is_green(c):
                    cells.add(c.coordinate)
        if cells:
            found[sn] = cells
    return found


def build_mapping():
    docs_out = {}
    summary = []
    for doc_key, d in DOCS.items():
        tpath = os.path.join(TPL, d["template"])
        greens = scan_green(tpath)
        curated = build_curated(doc_key)
        curated_index = {(e["sheet"], e["cell"]) for e in curated}

        entries = list(curated)
        # 緑セルのうちキュレーションに無いものを unknown で追加（要マッピングの可視化）
        unmapped = 0
        for sn, cells in greens.items():
            for cell in sorted(cells):
                if (sn, cell) not in curated_index:
                    entries.append(dict(sheet=sn, cell=cell, var="未割当_%s_%s" % (sn[:6], cell),
                                        source={"t": "unknown"}, tf="text",
                                        note="緑セルだが転記元未定義（要マッピング）"))
                    unmapped += 1
        # キュレーション済みだが緑で無いセル（2種の未マーク書類など）は残す（転記対象）
        n_curated_green = sum(1 for e in curated if e["sheet"] in greens and e["cell"] in greens[e["sheet"]])
        docs_out[doc_key] = dict(template=d["template"], entries=entries)
        summary.append((doc_key, len(curated), n_curated_green, unmapped,
                        sum(len(v) for v in greens.values())))

    # ----- SOP（17シート保持。緑マーク無し→キュレーションで直接指定）-----
    sop_path = os.path.join(TPL, SOP_TEMPLATE)
    if os.path.exists(sop_path):
        sop_entries = [dict(sheet=s, cell=c, var=v, source=src, tf=tf, note=note)
                       for (s, c, v, src, tf, note) in SOP_FIELDS]
        docs_out["SOP"] = dict(template=SOP_TEMPLATE, entries=sop_entries)
        summary.append(("SOP", len(sop_entries), 0, 0, 0))

    mapping = {
        "_about": "PRP申請書類 自動転記ツールのマッピング定義。transcribe.pyが読み込む。緑セル→ヒアリング/クロスシートの対応表。",
        "hearing_sheet": HEARING_SHEET,
        "documents": docs_out,
    }
    os.makedirs(MAP_DIR, exist_ok=True)
    out_path = os.path.join(MAP_DIR, "mapping.json")
    with io.open(out_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)
    print("mapping.json 生成:", out_path)
    for doc_key, n_cur, n_cur_green, n_unmap, n_green in summary:
        print("  [%s] curated=%d (うち緑一致=%d) / 緑総数=%d / 未割当(unknown)=%d"
              % (doc_key, n_cur, n_cur_green, n_green, n_unmap))
    return out_path


if __name__ == "__main__":
    build_mapping()
    print("Phase1 完了（mapping.json）")
