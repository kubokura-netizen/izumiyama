# -*- coding: utf-8 -*-
"""
e-再生医療（https://saiseiiryo.mhlw.go.jp/…）等のWebフォームへ、ヒアリングシートの値を
半自動で入力する補助ツール（方式B：ブラウザ起動→手動ログイン→自動入力→人が確認・送信）。

※ 送信は行いません。入力（下書き）までを自動化し、最終確認と送信は必ず人が行ってください。
※ 政府ポータルの自動操作は利用規約の確認が前提です。ログインは人が手動で行います。

使い方:
  1) 準備（初回のみ）:
       pip install playwright
       playwright install chromium
  2) フォーム項目の抽出（対応表を作るため）:
       py 99_data/src/web_fill.py --dump
     → ブラウザが開くのでログインし、対象フォームを表示 → コンソールでEnter
     → 03_logs/web_fields_dump.txt に入力欄の一覧（selector候補・ラベル）が出力される
  3) 対応表を作る:
       03_logs/web_fields_dump.txt を見ながら 99_data/マッピング/web_mapping.json の
       fields[].selector と source を埋める（source書式は docs_config と同じ）
  4) 自動入力:
       py 99_data/src/web_fill.py
     → ログイン→フォーム表示→Enter で各欄に自動入力（緑枠でハイライト）→ 人が確認して送信
"""
import os, sys, io, json, glob, datetime

SRC_DIR = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.dirname(SRC_DIR)
BASE = os.path.dirname(DATA)
sys.path.insert(0, SRC_DIR)

import transcribe as TX   # Hearing / resolve を再利用

MAPPING = os.path.join(DATA, "マッピング", "web_mapping.json")
PROFILE = os.path.join(DATA, "_web_profile")          # ログイン維持用プロファイル
LOGDIR = TX.resolve_dir("03_logs", create=True)


def find_hearing():
    """01_input の最新ヒアリング(.xlsx)を返す（web_fill 独自。--flag を誤検出しない）。"""
    d = TX.resolve_dir("01_input")
    cands = [c for c in glob.glob(os.path.join(d, "*.xlsx"))
             if not os.path.basename(c).startswith("~$")]
    cands.sort(key=os.path.getmtime, reverse=True)
    if not cands:
        raise FileNotFoundError("01_input にヒアリングシート(.xlsx)を入れてください")
    return cands[0]


def load_mapping():
    if not os.path.exists(MAPPING):
        return {}
    with io.open(MAPPING, encoding="utf-8") as f:
        return json.load(f)


# フォーム上の全入力欄を抽出するJS（name/id/type/ラベル/選択肢）
# ラジオ/チェックは name でグルーピングし、各選択肢のラベル文言（有/無・該当/非該当等）も収集。
DUMP_JS = r"""
() => {
  const optText = (el) => {
    // ラジオ/チェックの選択肢文言（隣接ラベル or 直後テキスト）
    if (el.id) { const l = document.querySelector('label[for="'+CSS.escape(el.id)+'"]'); if (l) return l.innerText.trim(); }
    let p = el.closest('label'); if (p) return p.innerText.trim();
    let sib = el.nextElementSibling; if (sib && sib.innerText) return sib.innerText.trim().slice(0,20);
    if (el.value) return el.value;
    return '';
  };
  const questionText = (el) => {
    // 設問ラベル（表の左セル/直近の見出し的テキスト）
    let td = el.closest('td,th'); if (td && td.previousElementSibling) return (td.previousElementSibling.innerText||'').trim().slice(0,60);
    let row = el.closest('tr,div'); if (row) { const t=(row.innerText||'').trim().split('\n')[0]; if (t) return t.slice(0,60); }
    if (el.getAttribute('aria-label')) return el.getAttribute('aria-label').trim();
    return '';
  };
  const sel = (el) => {
    if (el.id) return '#'+CSS.escape(el.id);
    if (el.name) return el.tagName.toLowerCase()+'[type="'+(el.type||'')+'"][name="'+el.name+'"]';
    return el.tagName.toLowerCase();
  };
  const seenGroup = {};
  const out = [];
  document.querySelectorAll('input, select, textarea').forEach(el => {
    const t = (el.type||el.tagName).toLowerCase();
    if (['hidden','submit','button','image','reset'].includes(t)) return;
    if (t==='radio' || t==='checkbox') {
      const key = t+':'+(el.name||el.id);
      if (seenGroup[key]) { seenGroup[key].choices.push(optText(el)); return; }
      const rec = {tag:'input', type:t, name:el.name||'', id:el.id||'',
                   selector: el.name? 'input[type="'+t+'"][name="'+el.name+'"]' : sel(el),
                   label: questionText(el), placeholder:'', choices:[optText(el)]};
      seenGroup[key]=rec; out.push(rec); return;
    }
    let opts = '';
    if (el.tagName.toLowerCase()==='select') opts = Array.from(el.options).map(o=>o.text.trim()).join(' | ');
    out.push({tag: el.tagName.toLowerCase(), type: t, name: el.name||'', id: el.id||'',
              selector: sel(el), label: questionText(el), placeholder: el.placeholder||'', options: opts, choices:[]});
  });
  return out;
}
"""


# チェック/ラジオ/ファイル欄/タブの“構造”を診断するJS（押下対象を特定するため）
DIAG_JS = r"""
() => {
  const vis = (el) => !!(el && (el.offsetWidth||el.offsetHeight||el.getClientRects().length));
  const cut = (s,n) => (s||'').replace(/\s+/g,' ').trim().slice(0,n);
  const sel = (el) => el.id ? '#'+el.id : (el.name ? el.tagName.toLowerCase()+'[name="'+el.name+'"]' : el.tagName.toLowerCase());
  const out = { checks: [], files: [], tabs: [] };
  // checkbox / radio の構造
  document.querySelectorAll('input[type="checkbox"], input[type="radio"]').forEach(el => {
    const id = el.id||'';
    const labFor = id ? document.querySelector('label[for="'+CSS.escape(id)+'"]') : null;
    let anc = el.closest('label');
    const cs = getComputedStyle(el);
    out.checks.push({
      type: el.type, name: el.name||'', id: id, selector: sel(el),
      checked: el.checked, inputVisible: vis(el), display: cs.display, opacity: cs.opacity,
      labelFor: !!labFor, labelForVisible: labFor ? vis(labFor) : false,
      labelForText: labFor ? cut(labFor.innerText, 30) : '',
      ancestorLabel: !!anc, ancestorLabelVisible: anc ? vis(anc) : false,
      parentTag: el.parentElement ? el.parentElement.tagName.toLowerCase() : '',
      parentClass: el.parentElement ? cut(el.parentElement.className, 40) : '',
      outer: cut(el.outerHTML, 160),
      parentOuter: el.parentElement ? cut(el.parentElement.outerHTML, 240) : ''
    });
  });
  // ファイル入力（非表示含む）＋アップロードらしき要素
  document.querySelectorAll('input[type="file"]').forEach(el => {
    out.files.push({ selector: sel(el), name: el.name||'', id: el.id||'', visible: vis(el),
                     accept: el.accept||'', multiple: !!el.multiple, outer: cut(el.outerHTML,160),
                     nearText: cut((el.closest('tr,li,div')||el).innerText, 60) });
  });
  document.querySelectorAll('button,a,label,span,div').forEach(el => {
    const t = (el.innerText||'').trim();
    if (/ファイル|アップロード|添付|参照|選択/.test(t) && t.length<20 && vis(el)) {
      out.files.push({ selector: sel(el), name:'(clickable)', id: el.id||'', visible:true,
                       accept:'', multiple:false, outer: cut(el.outerHTML,120), nearText: t });
    }
  });
  // タブらしき要素
  document.querySelectorAll('[role="tab"], .tab, li, a, button').forEach(el => {
    const t = (el.innerText||'').trim();
    if (/^(申請者情報|項目[1-7]|添付書類)$/.test(t) && vis(el)) {
      out.tabs.push({ tag: el.tagName.toLowerCase(), role: el.getAttribute('role')||'', text: t,
                      cls: cut(el.className,40), id: el.id||'', outer: cut(el.outerHTML,140) });
    }
  });
  return out;
}
"""


def dump_fields(page):
    fields = page.evaluate(DUMP_JS)
    diag = {}
    try:
        diag = page.evaluate(DIAG_JS)
    except Exception as e:
        diag = {"error": repr(e)}
    path = os.path.join(LOGDIR, "web_fields_dump.txt")
    with io.open(path, "w", encoding="utf-8") as f:
        f.write("Webフォーム 入力欄ダンプ  (%s)\n" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        f.write("URL: %s\n" % page.url)
        f.write("この selector と label を見て web_mapping.json を作成してください。\n\n")
        for i, fl in enumerate(fields, 1):
            f.write("%2d) [%s/%s] selector=%s\n" % (i, fl["tag"], fl["type"], fl["selector"]))
            if fl.get("label"):       f.write("     設問: %s\n" % fl["label"])
            if fl.get("placeholder"): f.write("     placeholder: %s\n" % fl["placeholder"])
            if fl.get("options"):     f.write("     選択肢(select): %s\n" % fl["options"])
            if fl.get("choices"):     f.write("     選択肢(radio/check): %s\n" % " / ".join([x for x in fl["choices"] if x]))
        # ---- 診断セクション ----
        f.write("\n\n================ 診断（チェック/ラジオ/ファイル/タブの構造）================\n")
        f.write("【checkbox/radio 構造】押下対象の特定用\n")
        for c in diag.get("checks", []):
            f.write("- %s [%s] name=%s checked=%s\n" % (c.get("selector"), c.get("type"), c.get("name"), c.get("checked")))
            f.write("    input可視=%s display=%s opacity=%s / labelFor=%s(可視%s)「%s」 / 祖先label=%s(可視%s)\n" % (
                c.get("inputVisible"), c.get("display"), c.get("opacity"),
                c.get("labelFor"), c.get("labelForVisible"), c.get("labelForText"),
                c.get("ancestorLabel"), c.get("ancestorLabelVisible")))
            f.write("    parent<%s class='%s'>\n" % (c.get("parentTag"), c.get("parentClass")))
            f.write("    input outer: %s\n" % c.get("outer"))
            f.write("    parent outer: %s\n" % c.get("parentOuter"))
        f.write("\n【ファイル入力/アップロード要素】\n")
        if not diag.get("files"):
            f.write("  （このタブでは検出なし。添付書類タブを表示してから再ダンプしてください）\n")
        for x in diag.get("files", []):
            f.write("- %s visible=%s multiple=%s accept=%s near「%s」\n     %s\n" % (
                x.get("selector"), x.get("visible"), x.get("multiple"), x.get("accept"), x.get("nearText"), x.get("outer")))
        f.write("\n【タブ要素】\n")
        for t in diag.get("tabs", []):
            f.write("- <%s role='%s' class='%s' id='%s'> 「%s」\n     %s\n" % (
                t.get("tag"), t.get("role"), t.get("cls"), t.get("id"), t.get("text"), t.get("outer")))
    print("→ 入力欄 %d件＋診断を出力: %s" % (len(fields), path))


import re as _re


def _pref_split(addr):
    """住所文字列を (都道府県, 残り) に分割。"""
    addr = TX.clean(addr)
    m = _re.match(r"(北海道|東京都|京都府|大阪府|.{1,3}?[県府])", addr)
    pref = m.group(1) if m else ""
    return pref, (addr[len(pref):] if pref else addr)


def load_output_ctx(mapping):
    """02_output の最新の出力フォルダ（output_folder_contains で絞る）を探し、
       様式xlsxシート ws とフォルダパス folder を返す。Webの参照元＝アウトプット。
       無ければ (None, '')。"""
    import openpyxl
    outdir = TX.resolve_dir("02_output")
    want = mapping.get("output_folder_contains", "2種関節系PRP")
    folders = [d for d in glob.glob(os.path.join(outdir, "*"))
               if os.path.isdir(d) and want in os.path.basename(d)]
    if not folders:
        return None, ""
    folder = max(folders, key=os.path.getmtime)
    ws = None
    xls = glob.glob(os.path.join(folder, "01.*.xlsx"))
    if xls:
        wb = openpyxl.load_workbook(xls[0], data_only=True)
        sh = mapping.get("output_sheet", "1 の２提供計画（治療）")
        ws = wb[sh] if sh in wb.sheetnames else wb[wb.sheetnames[0]]
    return ws, folder


def _iter_block_texts(doc):
    """docxを文書順（段落＋表）で走査し、各ブロックのテキストを返す。表は行を改行/セルを｜で連結。"""
    try:
        from docx.text.paragraph import Paragraph
        from docx.table import Table
    except Exception:
        for p in doc.paragraphs:
            yield p.text
        return
    for child in doc.element.body.iterchildren():
        tag = child.tag
        if tag.endswith("}p"):
            yield Paragraph(child, doc).text
        elif tag.endswith("}tbl"):
            tbl = Table(child, doc)
            rows = []
            for row in tbl.rows:
                cells = [c.text.strip() for c in row.cells]
                rows.append(" | ".join([c for c in cells if c]))
            yield "\n".join([r for r in rows if r.strip()])


def _docx_section(out_folder, file_contains, heading, until=None):
    """出力フォルダ内のWord（file_containsで特定）から、heading の次ブロック〜until手前 の本文を抽出。
       段落だけでなく表も文書順で拾う（適格性基準など表本文も取りこぼさない）。"""
    try:
        from docx import Document
    except Exception:
        return ""
    files = [f for f in glob.glob(os.path.join(out_folder, "*.docx"))
             if file_contains and file_contains in os.path.basename(f)]
    if not files:
        return ""
    doc = Document(files[0])
    blocks = list(_iter_block_texts(doc))
    start = None
    for i, t in enumerate(blocks):
        if heading and heading in t:
            start = i + 1
            break
    if start is None:
        return ""
    body = []
    for t in blocks[start:]:
        if until and t.strip().startswith(until):
            break
        body.append(t)
    return "\n".join(x for x in body if x.strip()).strip()


def _read_cells(ws, spec):
    """様式xlsxから値を読む。spec は次のいずれか：
         "L130"            … 単一セル
         "L130:L135"       … セル範囲（非空セルを文書順に改行連結）
         ["L130","L131"]   … 複数セル指定（順に改行連結）
       ★ Web長文欄は複数セルにまたがることがあるため、範囲/複数指定で結合できる。"""
    specs = spec if isinstance(spec, list) else [spec]
    parts = []
    for sp in specs:
        sp = str(sp).strip()
        if not sp:
            continue
        if ":" in sp:                                  # 範囲
            for row in ws[sp]:
                for c in row:
                    v = TX.clean(c.value)
                    if v:
                        parts.append(v)
        else:                                          # 単一
            v = TX.clean(ws[sp].value)
            if v:
                parts.append(v)
    return "\n".join(parts)


def _parse_date_parts(raw):
    """日付セルを (年, 月, 日) に解釈する。Excelの書き方の違いを吸収する：
         ・datetime/date          （日付書式のセル）
         ・「2026年 6月 15日」「2026/6/15」（文字列）
         ・46220 のようなシリアル値（日付書式が付いていないと数値のまま入る）
       未確定（XXXX年 等）や解釈不能なら None を返す（→ 欄は空のまま）。"""
    if raw is None:
        return None
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return (raw.year, raw.month, raw.day)

    def _from_serial(n):
        try:
            n = float(n)
        except Exception:
            return None
        if not (20000 <= n <= 80000):                    # 1954年頃〜2119年頃のみ日付とみなす
            return None
        d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(n))
        return (d.year, d.month, d.day)

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return _from_serial(raw)
    s = TX.clean(raw)
    if not s:
        return None
    mt = (_re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", s)
          or _re.search(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", s))
    if mt:
        return (int(mt.group(1)), int(mt.group(2)), int(mt.group(3)))
    if _re.fullmatch(r"\d{5}(\.\d+)?", s):               # 文字列化されたシリアル値
        return _from_serial(s)
    return None


def _date_part(raw, tf):
    """_parse_date_parts の結果を Web の選択肢に合わせて返す（年=2026 / 月=07 / 日=17）。"""
    p = _parse_date_parts(raw)
    if not p:
        return ""
    y, m, d = p
    return {"date_year": str(y), "date_month": "%02d" % m, "date_day": str(d)}.get(tf, "")


def _find_heading_row(ws, heading, occ=1, exact=False, head_cols=("B", "C", "D")):
    """見出し文字列に一致する行番号を返す。無ければ None。
         occ   … 同じ見出しが複数ある時、何個目か（実施責任者/事務担当者の「氏名」等）
         exact … True で完全一致（「所属機関」が「所属機関の郵便番号」に誤ヒットするのを防ぐ）"""
    if not heading or ws is None:
        return None
    n = 0
    for r in range(1, ws.max_row + 1):
        for hc in head_cols:
            t = TX.clean(ws["%s%d" % (hc, r)].value)
            if not t:
                continue
            hit = (t == heading) if exact else (heading in t)
            if hit:
                n += 1
                if n >= occ:
                    return r
                break
    return None


def _read_by_heading(ws, heading, col="L", max_span=40, skip=0, occ=1,
                     exact=False, offset=0, span=0, head_cols=("B", "C", "D"),
                     joiner="\n\n"):
    """★見出し文字列で行を探して値を読む（行番号のハードコードを避ける）。
       転記ツール(transcribe.py)の改修で出力Excelの行がズレても自動追従できる。
         skip/offset … 見出し行から読み飛ばす行数（見出し行が■/□マーカーの時は1）
         span        … 読む行数を固定（0=次の見出しが来るまで。医師の氏名/所属など1行だけ取る用）
         occ / exact … 重複見出しの選別
         joiner      … 複数セルを連結する区切り。既定は空行 "\\n\\n"（＝セル境界を1行空けて
                        本文中の改行と区別する）。1行詰めたい欄は row.joiner:"\\n" で上書き。
       戻り値: 連結テキスト（見出しが見つからなければ ""）。"""
    start = _find_heading_row(ws, heading, occ, exact, head_cols)
    if not start:
        return ""
    # offset は負も可（値が見出しの“上”にある様式ヘッダ部に対応。例: R11=値 / R12=「名　称」）
    begin = start + (offset if offset else skip)
    if begin < 1:
        return ""
    stop = (begin + span) if span > 0 else min(start + max_span, ws.max_row + 1)
    parts = []
    for r in range(max(begin, 1), min(stop, ws.max_row + 1)):
        if span <= 0 and r > start and \
           any(TX.clean(ws["%s%d" % (hc, r)].value) for hc in head_cols):
            break                                        # 次の見出しが来たら終了
        v = TX.clean(ws["%s%d" % (col, r)].value)
        if v:
            parts.append(v)
    return joiner.join(parts)


def _resolve_value(fld, hearing, kits, out_ws=None, out_folder=""):
    """フィールドの入力値を決める。★出力(アウトプット)を最優先：
       cell/cells=様式xlsxのセル（単一/範囲/複数） / docx=出力Wordの見出しセクション。
       無ければヒアリング(source)。web専用type: pref/addr_body/today。cell_tf: pref/addr_body/zip。"""
    # ⓪ choice_from: Excelの■/□マーカーから選択肢を決める（補償の有無 有/無 等）
    #    書式: "choice_from": { "options": [["有","L148"],["無","R148"]], "marked": "■" }
    #    → 各[ラベル,セル]を見て、セル値に marked(既定■) を含む最初のラベルを返す。
    cf = fld.get("choice_from")
    if cf and out_ws is not None:
        marked = cf.get("marked", "■")
        # heading指定なら見出しで行を特定し、options の第2要素を「列文字」として扱う
        # （例: {"heading":"補償の有無","options":[["有","L"],["無","R"]]}）。
        # heading無しなら従来どおり第2要素はセル番地（例 "L148"）。
        hrow = None
        if cf.get("heading"):
            hrow = _find_heading_row(out_ws, cf["heading"], int(cf.get("occ", 1)))
            if not hrow:
                return ""
        for opt in cf.get("options", []):
            if not (isinstance(opt, (list, tuple)) and len(opt) >= 2):
                continue
            label, cref = opt[0], opt[1]
            ref = ("%s%d" % (cref, hrow)) if hrow else cref
            cv = TX.clean(out_ws[ref].value)
            if cv and marked in cv:
                return label

    # ①-A ★見出しで行を探して読む（行番号ハードコードより優先。レイアウト変更に強い）
    #      書式: "row": { "heading": "細胞提供者の選定方法", "col": "L" }
    rowspec = fld.get("row")
    if isinstance(rowspec, dict) and out_ws is not None:
        hcols = tuple(rowspec.get("head_cols") or ("B", "C", "D"))   # 見出しを探す列
        if rowspec.get("tf") in ("date_year", "date_month", "date_day"):
            # ★日付はセルの生値を読む（TX.cleanで文字列化するとシリアル値/日付型の情報が落ちるため）
            hr = _find_heading_row(out_ws, rowspec.get("heading", ""),
                                   int(rowspec.get("occ", 1)),
                                   bool(rowspec.get("exact", False)), hcols)
            if not hr:
                return ""
            r0 = hr + int(rowspec.get("offset", rowspec.get("skip", 0)))
            return _date_part(out_ws["%s%d" % (rowspec.get("col", "L"), r0)].value,
                              rowspec["tf"])
        v = _read_by_heading(out_ws, rowspec.get("heading", ""),
                             rowspec.get("col", "L"),
                             int(rowspec.get("max_span", 40)),
                             int(rowspec.get("skip", 0)),
                             int(rowspec.get("occ", 1)),
                             bool(rowspec.get("exact", False)),
                             int(rowspec.get("offset", 0)),
                             int(rowspec.get("span", 0)),
                             hcols,
                             rowspec.get("joiner", "\n\n"))   # 既定=空行でセル境界を区切る
        # 住所→都道府県/以降、郵便番号の整形（cell経路と同じ変換を row でも使えるように）
        tf = rowspec.get("tf")
        if tf in ("pref", "addr_body"):
            pref, body = _pref_split(v)
            if pref or body:
                return pref if tf == "pref" else body
        elif tf == "zip":
            z = (v or "").lstrip("〒 　").strip()
            if z:
                return z
        if rowspec.get("tf") in ("date_year", "date_month", "date_day") and v:
            mt = (_re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", v)
                  or _re.search(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", v))
            if not mt:
                return ""                                # 日付未確定（XXXX年 等）→空
            y, m, d = int(mt.group(1)), int(mt.group(2)), int(mt.group(3))
            return {"date_year": str(y), "date_month": "%02d" % m,
                    "date_day": str(d)}[rowspec["tf"]]
        if v:
            return v

    # ①-B アウトプット様式xlsxのセル優先（単一/範囲/複数セルの結合に対応）
    cell = fld.get("cell") or fld.get("cells")
    if cell and out_ws is not None:
        tf = fld.get("cell_tf")
        if tf in ("date_year", "date_month", "date_day"):
            # 日付セルを 年/月/日 に分解（Web側が年月日の3プルダウンのため）。
            # datetime / 「2026年 6月 15日」/ シリアル値(46220) のどれでも解釈する。
            first = cell[0] if isinstance(cell, list) else str(cell).split(":")[0]
            return _date_part(out_ws[first].value, tf)
        if tf in ("pref", "addr_body", "zip"):
            # 住所/郵便番号の分割は単一セル前提（範囲指定時は先頭セルを使用）
            first = cell[0] if isinstance(cell, list) else str(cell).split(":")[0]
            raw = TX.clean(out_ws[first].value)
            if tf in ("pref", "addr_body"):
                pref, body = _pref_split(raw)
                if pref or body:
                    return pref if tf == "pref" else body
            else:  # zip
                z = raw.lstrip("〒 　").strip()
                if z:
                    return z
        else:
            v = _read_cells(out_ws, cell)
            if v:
                return v
    # ② アウトプットWordのセクション
    dx = fld.get("docx")
    if dx and out_folder:
        v = TX.clean(_docx_section(out_folder, dx.get("file_contains", ""),
                                   dx.get("heading", ""), dx.get("until")))
        if v:
            return v
    # ③ ヒアリング（フォールバック）
    src = fld.get("source")
    if src:
        t = src.get("t")
        if t in ("pref", "addr_body"):
            inner = src.get("source") or {"t": "hearing",
                    "label": "医療機関/住所（診療所開設届上）", "section": "法人/医療機関"}
            addr, _ = TX.resolve(inner, hearing, kits)
            pref, body = _pref_split(addr)
            return pref if t == "pref" else body
        if t == "today" and src.get("fmt") in ("year", "month", "day"):
            # ※ここで import datetime すると関数全体で datetime がローカル扱いになり、
            #   上の date_year/month/day 分解で UnboundLocalError になる（モジュール先頭で import 済み）。
            d = datetime.datetime.now()
            return {"year": str(d.year), "month": "%02d" % d.month, "day": str(d.day)}[src.get("fmt")]
        v, _ = TX.resolve(src, hearing, kits)
        v = TX.clean(v)
        # トークン名指定（採血量・治療価格・キット製造方法等の汎用解決）
        if not v and t == "token":
            nm = src.get("name", "")
            v = TX.clean(TX.resolve_token_by_name(nm, hearing, kits, fld.get("saisei", 1)))
            if not v and nm in ("採取方法", "加工方法", "投与方法"):   # 採取の非整合を補完
                part = {"採取方法": "採取", "加工方法": "加工", "投与方法": "投与"}[nm]
                v = TX.clean(hearing.kit_block(part, numbered=True))
        return v
    return TX.clean(fld.get("value", ""))


def _locate(page, fld):
    """selector優先、無ければ label（アクセシブルラベル）で要素を特定。"""
    selector = (fld.get("selector") or "").strip()
    if selector:
        return page.locator(selector).first
    label = (fld.get("label") or "").strip()
    if label:
        try:
            loc = page.get_by_label(label, exact=False).first
            if loc.count() > 0:
                return loc
        except Exception:
            pass
        return page.get_by_text(label, exact=False).first
    return None


def fill_one_page(page, mapping, hearing, kits, filled_keys, out_ws=None, out_folder=""):
    """現在表示中のページの、まだ埋めていないフィールドを入力する。"""
    done = 0
    logs = []
    for idx, fld in enumerate(mapping.get("fields", [])):
        if not isinstance(fld, dict):
            continue                              # 説明用の文字列などはスキップ
        key = fld.get("desc", "") + "#" + str(idx)
        if key in filled_keys:
            continue
        if not (fld.get("selector") or fld.get("label")):
            continue
        val = _resolve_value(fld, hearing, kits, out_ws, out_folder)
        desc = fld.get("desc", fld.get("selector") or fld.get("label"))
        ftype = (fld.get("type") or "text").lower()
        try:
            if ftype == "radio":
                # 選ぶ選択肢ラベル＝ choice(固定) or 値そのもの（有/無・該当/非該当）
                choice = fld.get("choice") or val
                grp = (fld.get("selector") or "").strip()
                if not choice or not grp:
                    continue
                # グループ内(name一致)の各ラジオのラベル文言を見て、choiceに一致するものだけをチェック
                radios = page.locator(grp)
                cnt = radios.count()
                picked = False
                for i in range(cnt):
                    r = radios.nth(i)
                    lbl = r.evaluate(
                        "el => { if(el.id){const l=document.querySelector('label[for=\"'+el.id+'\"]'); if(l) return l.innerText.trim();}"
                        " let p=el.closest('label'); if(p) return p.innerText.trim();"
                        " let s=el.nextElementSibling; if(s&&s.innerText) return s.innerText.trim();"
                        " return el.value||''; }")
                    if choice and (choice == (lbl or "").strip() or choice in (lbl or "")):
                        r.check()
                        r.evaluate("el => { const l=el.closest('label')||el.parentElement; if(l){l.style.outline='2px solid #00B050';} }")
                        picked = True
                        break
                if not picked:
                    try:
                        page.locator("%s[value='%s']" % (grp, choice)).first.check(); picked = True
                    except Exception:
                        pass
                if picked:
                    logs.append("選択(radio): %s → %s" % (desc, choice))
                    filled_keys.add(key); done += 1
                else:
                    logs.append("radio未選択(候補なし): %s → %s" % (desc, choice))
                continue
            loc = _locate(page, fld)
            if loc is None or loc.count() == 0:
                continue                         # このページには無い（別セクションかも）→スキップ
            if not val:
                continue
            tag = loc.evaluate("el => el.tagName.toLowerCase()")
            if tag == "select" or ftype == "select":
                try:
                    loc.select_option(label=val)
                except Exception:
                    loc.select_option(value=val)
            elif ftype == "checkbox":
                (loc.check() if str(val) not in ("", "0", "false", "無", "×") else loc.uncheck())
            else:
                loc.fill(str(val))
            loc.evaluate("el => { el.style.outline='2px solid #00B050'; el.style.background='#eaffea'; }")
            logs.append("入力: %s = %s" % (desc, str(val)[:40]))
            filled_keys.add(key); done += 1
        except Exception as e:
            logs.append("失敗: %s %r" % (desc, e))
    for l in logs:
        print("  -", l)
    return done


def fill_fields(page, mapping, hearing, kits):
    """複数セクション対応：現ページを入力→ユーザーが次セクションへ→また入力…を繰り返す。"""
    out_ws, out_folder = load_output_ctx(mapping)
    if out_folder:
        print("参照元(アウトプット): %s ／ 様式xlsx=%s"
              % (os.path.basename(out_folder), "あり" if out_ws is not None else "なし"))
    else:
        print("※ アウトプット未検出 → ヒアリングから取得します（先に 転記実行.bat を実行推奨）")
    filled = set()
    total = 0
    while True:
        print("\n=== このページを自動入力 ===")
        n = fill_one_page(page, mapping, hearing, kits, filled, out_ws, out_folder)
        total += n
        print("  （このページで %d件入力・累計 %d件）" % (n, total))
        print("★ 内容を確認してください。ツールは送信しません。")
        ans = input("次のセクションを表示したら Enter（続けて入力）／ 終了は q + Enter : ").strip().lower()
        if ans == "q":
            break
    print("\n合計 %d件を入力しました。最終確認のうえ、画面から送信してください。" % total)


# ============================================================
#  自動一括モード（--auto）：全タブを自動送り→全欄入力→一時保存→不備レポート
#   ・送信/申請/提出/確定 は絶対に押しません（一時保存のみ許可）。
#   ・不備で止まった場合、どの欄・どのメッセージで止まったかを報告します。
# ============================================================

# ラジオの選択肢ラベル文言を取るJS（fill_one_page と同じ規則）
_RADIO_LABEL_JS = (
    "el => { if(el.id){const l=document.querySelector('label[for=\"'+el.id+'\"]'); if(l) return l.innerText.trim();}"
    " let p=el.closest('label'); if(p) return p.innerText.trim();"
    " let s=el.nextElementSibling; if(s&&s.innerText) return s.innerText.trim();"
    " return el.value||''; }")
# チェック/ラジオを確実に選択：クリック（トグル）せず checked を直接セットし、input/change を発火。
# ※ inputのクリックがlabelハンドラに伝播して二重トグル→差し引きゼロになる問題を避けるため。
_SET_CHECKED_JS = ("(e, v) => { e.checked = v;"
                   " e.dispatchEvent(new Event('input',  {bubbles:true}));"
                   " e.dispatchEvent(new Event('change', {bubbles:true})); }")
_HILITE_JS = "el => { const l=el.closest('label')||el.parentElement; if(l){l.style.outline='2px solid #00B050';} }"
_HILITE_JS2 = "el => { el.style.outline='2px solid #00B050'; el.style.background='#eaffea'; }"

# 絶対に押してはいけない語（本申請の送信系）。一時保存の自動化から除外する安全ガード。
_FORBIDDEN_BTN = ("送信", "申請", "提出", "確定", "登録完了", "完了する")
# 一時保存（下書き保存）として許可するボタン文言。左優先。
_SAVE_NAMES = ("一時保存", "下書き保存", "下書き保存する", "一時保存する")
# 次のタブ/セクションへ進む操作の候補文言。
_NEXT_NAMES = ("次の項目へ", "次へ進む", "次のページへ", "次へ", "次のページ", "次")


def _field_display(fld, idx):
    return fld.get("desc") or fld.get("selector") or fld.get("label") or ("field#%d" % idx)


def _web_len(s):
    """Web側の文字数カウントに合わせる。改行は \\r\\n = 2文字として計上される前提で数える
       （表記は4000字でも、改行の多い本文は体感3200字程度で溢れるため）。"""
    return len(s) + s.count("\n")


def _apply_overflow(item, mapping):
    """②文字数オーバー対策：値が上限を超えたときの扱いを決める。
       上限= フィールドの char_limit → mapping.char_limit_default。未設定なら判定しない。
       mode= フィールドの overflow_mode → mapping.overflow_mode
         mark    …【既定】本文を入れず「文字数超過」と記入する。
                   ※長すぎる本文を入れるとサイトが弾いて一時保存まで到達できないため、
                     保存を必ず通し、あとから該当欄だけ手直しできるようにする。
         truncate… 末尾を「…」で切って上限内に収める
         report  … 値はそのまま入れて、超過をレポートするだけ（保存が止まる恐れあり）"""
    val = item.get("val")
    if not isinstance(val, str) or not val:
        return
    fld = item["fld"]
    limit = fld.get("char_limit", mapping.get("char_limit_default"))
    if not limit:
        return
    n = _web_len(val)                                  # 改行を2文字として計上
    item["len"] = n
    item["limit"] = limit
    warn = fld.get("char_warn", mapping.get("char_warn_default"))
    if warn and warn <= n <= limit:
        item["warn_len"] = True                        # 上限手前（改行増で溢れる恐れ）
    if n > limit:
        item["over"] = n - limit
        mode = fld.get("overflow_mode", mapping.get("overflow_mode", "mark"))
        if mode == "truncate":
            cut = max(0, limit - 1)
            item["val"] = val[:cut] + "…"              # 末尾を省略記号にして上限内へ
            item["truncated"] = True
        elif mode == "mark":
            item["val"] = fld.get("overflow_mark", mapping.get("overflow_mark", "文字数超過"))
            item["marked_over"] = True                 # 本文は入れず目印だけ入れる


def build_plan(mapping, hearing, kits, out_ws, out_folder):
    """全マッピング欄の入力値を先に解決し、状態付きの作業リストを作る。"""
    plan = []
    for idx, fld in enumerate(mapping.get("fields", [])):
        if not isinstance(fld, dict):
            continue                                   # 説明用の文字列はスキップ
        if not (fld.get("selector") or fld.get("label")):
            continue
        val = _resolve_value(fld, hearing, kits, out_ws, out_folder)
        item = {"idx": idx, "fld": fld, "val": val,
                "desc": _field_display(fld, idx), "status": "pending", "error": ""}
        _apply_overflow(item, mapping)
        plan.append(item)
    return plan


def _qualify_sel(selector, ftype):
    """★CakePHP系フォームは checkbox/radio の直前に「同じname」の <input type="hidden"> を出力する。
       type未指定のセレクタ（例 input[name='data[Plan][cells_type1]']）はその hidden を先に掴んでしまい、
       hidden にはラベルが無いため「クリック対象なし」でスキップされる（＝入らない原因）。
       → checkbox/radio は type を明示して“実体”だけを掴む。"""
    s = (selector or "").strip()
    if not s or ftype not in ("checkbox", "radio"):
        return s
    if "type=" in s.replace(" ", ""):
        return s                                          # 既にtype指定済み
    if s.startswith("input["):
        return "input[type='%s']%s" % (ftype, s[len("input"):])
    return s


def _visible_label(page, el, exact_selector=""):
    """checkbox/radio の“直近ラベル(closest)”を可視なら返す。無ければ None。
       ★重要: 「そのinputを内包するlabel」を locator の .first で取ると、DOM順で“外側の親label”を
       掴んでしまい、クリックが別の選択肢に当たる（＝丸は選ばれない／四角は二重トグルでOFF）。
       そこで JS の closest('label') で直近ラベルに一時マークを付け、それだけを確実に掴む。"""
    # ① 直近ラベル（closest）に一時マークを付けて取得＝最も確実
    try:
        ok = el.evaluate(
            "e => { const l = e.closest('label'); if (!l) return false;"
            " document.querySelectorAll('[data-pwlab]').forEach(x => x.removeAttribute('data-pwlab'));"
            " l.setAttribute('data-pwlab', '1'); return true; }")
        if ok:
            lab = page.locator("label[data-pwlab='1']").first
            if lab.count() > 0 and lab.is_visible():
                return lab
    except Exception:
        pass
    # ② label[for=id]
    idv = ""
    try:
        idv = el.get_attribute("id") or ""
    except Exception:
        idv = ""
    if idv:
        try:
            lab = page.locator("label[for=\"%s\"]" % idv).first
            if lab.count() > 0 and lab.is_visible():
                return lab
        except Exception:
            pass
    return None


def _click_to_state(page, el, want, exact_selector=""):
    """★可視ラベルを“実クリック”して checked を want にする（人の操作と同じ＝サイトのハンドラが発火）。
       checkedを直接代入するとサイトが認識しない（画面・検証が更新されない）ため必ずクリックする。
       戻り値: "ok" / "absent"(タブ未表示→後で) / "ng"(クリックしたが状態が合わない)。"""
    lab = _visible_label(page, el, exact_selector)
    if lab is None:
        return "absent"

    def _state():
        # ★JSで checked を直接読む。is_checked() は例外を投げることがあり、
        #   それを「未チェック」と誤判定して“もう一度クリック→OFFに戻す”事故になるため。
        try:
            return bool(el.evaluate("e => !!e.checked"))
        except Exception:
            return None

    if _state() == want:
        return "ok"
    try:
        lab.click()                                      # 直近ラベルを1回だけ実クリック（人と同じ）
    except Exception:
        return "ng"
    page.wait_for_timeout(400)                           # サイトのハンドラが状態を確定するまで待つ
    # ★追加クリックはしない：合わない場合は二重トグル等が起きている証拠なので、
    #   もう一度押しても戻るだけ（＝“入って消える”の原因）。正直に ng を返して報告する。
    return "ok" if _state() == want else "ng"


def _click_labeled_input(page, el):
    """checkbox/radio を、対応する“可視ラベル”を実クリックしてON（カスタム装飾UI対応）。
       ★input直接操作ではフレームワークのstateが更新されず反映されないため、実クリックする。
       優先: label[for=id] → 祖先label → input可視クリック。
       可視のクリック対象が無ければ False（＝別タブで未表示→呼び出し側で後回し）。"""
    idv = ""
    try:
        idv = el.get_attribute("id") or ""
    except Exception:
        idv = ""
    if idv:
        try:
            lab = page.locator("label[for=\"%s\"]" % idv).first
            if lab.count() > 0 and lab.is_visible():
                lab.click()
                return True
        except Exception:
            pass
    try:                                                # inputを内包する祖先label
        anc = el.locator("xpath=ancestor::label[1]").first
        if anc.count() > 0 and anc.is_visible():
            anc.click()
            return True
    except Exception:
        pass
    try:                                                # input自体（可視なら実クリック）
        if el.is_visible():
            el.click()
            return True
    except Exception:
        pass
    # ★ここでJSのinput.click()はしない：inputのクリックがlabelハンドラに伝播して
    #   チェックボックスが二重トグル（差し引きゼロ）になるため。可視ラベルが無い＝別タブなので
    #   Falseを返し、呼び出し側で「absent」→そのタブが開いた時に可視ラベルを実クリックさせる。
    return False


def _click_tab(page, name):
    """タブ見出し（項目1〜項目7 等）を名前で開く。開けたら True。"""
    for role in ("tab", "link", "button"):
        try:
            t = page.get_by_role(role, name=name, exact=True)
            n = t.count()
        except Exception:
            n = 0
        for i in range(n):
            try:
                cand = t.nth(i)
                if cand.is_visible():
                    cand.click()
                    return True
            except Exception:
                continue
    try:
        t = page.get_by_text(name, exact=True)
        for i in range(min(t.count(), 5)):
            cand = t.nth(i)
            if cand.is_visible():
                cand.click()
                return True
    except Exception:
        pass
    return False


def _place_field(page, item):
    """現在表示中のタブに item の欄があれば入力する。
       戻り値: placed / absent(このタブには無い) / empty(欄はあるが値が空) /
               no-choice(ラジオの選択肢不一致) / error(例外)。"""
    fld = item["fld"]
    val = item["val"]
    ftype = (fld.get("type") or "text").lower()
    try:
        if ftype == "radio":
            # ★type付きに正規化（同名hidden inputを掴まないように）
            grp = _qualify_sel(fld.get("selector"), "radio")
            if not grp:
                return "empty"
            radios = page.locator(grp)
            cnt = radios.count()
            if cnt == 0:
                return "absent"                        # このDOMに無い
            choice = fld.get("choice") or val
            if not choice:
                return "empty"
            # 対象ラジオを選ぶ（ラベル文言一致 → value一致）
            target = None
            for i in range(cnt):
                r = radios.nth(i)
                try:
                    lbl = r.evaluate(_RADIO_LABEL_JS)
                except Exception:
                    lbl = ""
                if choice == (lbl or "").strip() or choice in (lbl or ""):
                    target = r
                    break
            if target is None:
                try:
                    cand = page.locator("%s[value='%s']" % (grp, choice)).first
                    if cand.count() > 0:
                        target = cand
                except Exception:
                    pass
            if target is None:
                return "no-choice"
            # 対象ラジオを1つに絞れるCSS（value属性）を作り、それを内包するlabelを確実に取得する
            exact = ""
            try:
                vattr = target.get_attribute("value") or ""
                if vattr:
                    exact = "%s[value='%s']" % (grp, vattr)
            except Exception:
                pass
            # ★可視ラベルを実クリック（タブ未表示なら absent＝そのタブを開いた時に再試行）
            st = _click_to_state(page, target, True, exact)
            if st == "absent":
                return "absent"
            if st == "ok":
                return "placed"
            item["error"] = "ラベルをクリックしたが選択状態にならない"
            return "error"

        if ftype == "checkbox":
            # ★type付きに正規化して掴む（同名hidden inputを掴むとラベルが無く操作不能になる）
            qsel = _qualify_sel(fld.get("selector"), "checkbox")
            loc = page.locator(qsel).first if qsel else _locate(page, fld)
        else:
            loc = _locate(page, fld)
        if loc is None or loc.count() == 0:
            return "absent"
        if not val:
            return "empty"
        tag = loc.evaluate("el => el.tagName.toLowerCase()")
        typ = ""
        try:
            typ = (loc.get_attribute("type") or "").lower()
        except Exception:
            pass
        if tag == "select" or ftype == "select":
            # ★可視でなければ後回し（別タブのselectに select_option すると可視待ちで数十秒固まる）
            try:
                if not loc.is_visible():
                    return "absent"
            except Exception:
                pass
            try:
                loc.select_option(label=val)
            except Exception:
                loc.select_option(value=val)
        elif ftype == "checkbox" or typ == "checkbox":
            # ★可視ラベルを実クリック（checkedの直接代入だとサイトが認識せず画面・検証が更新されない）
            want = str(val) not in ("", "0", "false", "無", "×")
            st = _click_to_state(page, loc.first, want, _qualify_sel(fld.get("selector"), "checkbox"))
            if st == "absent":
                return "absent"                        # タブ未表示→そのタブを開いた時に再試行
            if st == "ok":
                return "placed"
            item["error"] = "ラベルをクリックしたがチェック状態にならない"
            return "error"
        else:
            # テキスト/テキストエリアは可視でないと fill できない（別タブなら後で）
            try:
                if not loc.is_visible():
                    return "absent"
            except Exception:
                pass
            loc.fill(str(val))
        try:
            loc.evaluate(_HILITE_JS2)
        except Exception:
            pass
        return "placed"
    except Exception as e:
        item["error"] = repr(e)
        return "error"


def _btn_text(loc):
    try:
        return (loc.inner_text() or "").strip()
    except Exception:
        try:
            return (loc.get_attribute("value") or "").strip()
        except Exception:
            return ""


def _is_forbidden(text):
    return any(w in (text or "") for w in _FORBIDDEN_BTN)


def _find_clickable(page, names, explicit_selector=None):
    """names のいずれかの文言を持つ、押せる要素を探す。送信系は除外。
       戻り値: (locator or None, matched_text)。"""
    if explicit_selector:
        try:
            b = page.locator(explicit_selector).first
            if b.count() > 0 and b.is_enabled():
                t = _btn_text(b)
                if not _is_forbidden(t):
                    return b, t
        except Exception:
            pass
    for name in names:
        for role in ("button", "link"):
            try:
                b = page.get_by_role(role, name=name, exact=False)
                n = b.count()
            except Exception:
                n = 0
            for i in range(n):
                cand = b.nth(i)
                try:
                    if not cand.is_visible() or not cand.is_enabled():
                        continue
                except Exception:
                    continue
                t = _btn_text(cand) or name
                if _is_forbidden(t):
                    continue
                return cand, t
        # input[type=button|submit] を value で
        for typ in ("button", "submit"):
            try:
                b = page.locator("input[type='%s'][value*='%s']" % (typ, name))
                if b.count() > 0 and b.first.is_visible() and b.first.is_enabled():
                    t = _btn_text(b.first) or name
                    if not _is_forbidden(t):
                        return b.first, t
            except Exception:
                pass
    return None, ""


def _click_next(page, mapping):
    """次のタブ/セクションへ進む。進めたら True。最後のタブなら False。"""
    loc, _ = _find_clickable(page, _NEXT_NAMES, mapping.get("next_selector"))
    if loc is None:
        return False
    try:
        loc.click()
        page.wait_for_timeout(400)                     # 瞬時切替でもDOM反映待ち
        return True
    except Exception:
        return False


def _collect_site_errors(page, mapping):
    """サイトが表示する検証（不備）メッセージを収集。"""
    cands = []
    if mapping.get("error_selector"):
        cands.append(mapping["error_selector"])
    cands += [".error", ".errorMessage", ".error-message", ".help-block.error",
              ".text-danger", ".is-error", ".invalid-feedback",
              "[class*='error']:not(input):not(select):not(textarea)",
              ".alert-danger", "[role='alert']"]
    msgs = []
    for c in cands:
        try:
            loc = page.locator(c)
            for i in range(min(loc.count(), 80)):
                el = loc.nth(i)
                try:
                    if not el.is_visible():
                        continue
                    t = (el.inner_text() or "").strip()
                except Exception:
                    t = ""
                if t and len(t) < 300:
                    msgs.append(" ".join(t.split()))
        except Exception:
            pass
    # 重複除去（順序保持）
    seen, out = set(), []
    for m in msgs:
        if m not in seen:
            seen.add(m)
            out.append(m)
    return out


def _click_return_to_edit(page):
    """一時保存の確認ページから「保存データ編集に戻る」で下書きを開く（添付のため）。
       ※「入力内容確認」等の送信に向かう操作は押さない。"""
    loc, _text = _find_clickable(page, ("保存データ編集に戻る", "データ編集に戻る", "編集に戻る"))
    if loc is None:
        return False
    try:
        loc.click()
        # ★ページ遷移が完了するまで待つ（先走ると添付欄が未ロードのまま探しに行くため）
        for st in ("domcontentloaded", "load", "networkidle"):
            try:
                page.wait_for_load_state(st, timeout=15000)
            except Exception:
                pass
        page.wait_for_timeout(1500)
        return True
    except Exception:
        return False


def _extract_receipt(page):
    """一時保存後の確認ページから 受付番号 と パスワード を抽出する（将来のメール通知用）。
       戻り値: {"受付番号":..., "パスワード":...}（取れたものだけ）。"""
    info = {}
    try:
        txt = page.evaluate("() => document.body ? document.body.innerText : ''") or ""
    except Exception:
        txt = ""
    # ラベル直後（空白/改行のみ）の英数字を値とする。説明文「受付番号とパスワードで…」には
    # 助詞「と/で」が挟まるので誤マッチしない。
    for key in ("受付番号", "パスワード"):
        m = _re.search(key + r"[\s　:：]*([0-9A-Za-z][0-9A-Za-z\-]{5,})", txt)
        if m:
            info[key] = m.group(1)
    # 受付番号とパスワードが同値になった場合は取り違えの疑い→パスワードは別トークンを再探索
    if info.get("受付番号") and info.get("受付番号") == info.get("パスワード"):
        rest = txt.split("パスワード", 1)[-1]
        for m in _re.finditer(r"([0-9A-Za-z][0-9A-Za-z\-]{5,})", rest):
            if m.group(1) != info["受付番号"]:
                info["パスワード"] = m.group(1)
                break
    return info


def _click_save(page, mapping):
    """一時保存（下書き保存）ボタンをクリック。送信系は押さない。
       戻り値: (clicked:bool, text:str)。"""
    loc, text = _find_clickable(page, _SAVE_NAMES, mapping.get("save_selector"))
    if loc is None:
        return False, ""
    if _is_forbidden(text):
        return False, text                             # 念のため二重ガード
    try:
        loc.click()
        page.wait_for_timeout(1200)
        return True, text
    except Exception:
        return False, text


# 添付書類タブの各行（「ファイル選択」ボタン）と、その行見出し（先頭番号付き文書名）を実行時に発見するJS
ATTACH_ROWS_JS = r"""
() => {
  const cut = (s,n) => (s||'').replace(/\s+/g,' ').trim().slice(0,n);
  const btns = [...document.querySelectorAll('button,a,label')]
      .filter(e => /ファイル選択/.test(e.innerText||''));
  return btns.map((el,i) => {
    let row = el, head = '';
    for (let n=0; n<8 && row; n++) {
      const lines = (row.innerText||'').split('\n').map(s=>s.trim()).filter(Boolean);
      const hit = lines.find(L => /^\d+(\.\d+)?[ 　]/.test(L));   // 「2 提供…」「4.5 …」
      if (hit) { head = cut(hit, 60); break; }
      row = row.parentElement;
    }
    return { index:i, heading:head };
  });
}
"""


def _lead_num(name):
    m = _re.match(r"\s*([0-9]+(?:\.[0-9]+)?)", (name or "").strip())
    return m.group(1) if m else ""


def _manual_attach_table(out_folder):
    """添付を手動で行うための対応表（頭の数字＝Webのスロット番号）を返す。"""
    lines = []
    if not out_folder:
        return lines
    docs = [f for f in glob.glob(os.path.join(out_folder, "*.docx"))
            if not os.path.basename(f).startswith("~$")]
    nummap = {}
    for f in docs:
        n = _lead_num(os.path.basename(f))
        if n:
            nummap.setdefault(n, []).append(f)
    if not nummap:
        return lines
    lines.append("  ―― 手動アップロード対応表（添付書類タブで各スロットに）――")
    for n in sorted(nummap, key=lambda x: float(x) if x.replace('.', '', 1).isdigit() else 999):
        for f in sorted(nummap[n]):
            lines.append("    スロット%s ← %s" % (n, os.path.basename(f)))
    return lines


# 添付書類パネルの実DOMを診断するJS（--attach で構造を確認するため）
ATTACH_DIAG_JS = r"""
() => {
  const cut = (s,n) => (s||'').replace(/\s+/g,' ').trim().slice(0,n);
  const vis = (e) => !!(e && (e.offsetWidth || e.offsetHeight || e.getClientRects().length));
  let panel = document.querySelector('#tab9');
  if (!panel) {
    const t = [...document.querySelectorAll('[role=tab]')].find(e => /添付書類/.test(e.innerText||''));
    if (t) { const id = t.getAttribute('aria-controls'); if (id) panel = document.getElementById(id); }
  }
  const scope = panel || document;
  const clickables = [...scope.querySelectorAll('button,a,label,input[type=button],input[type=submit]')]
      .map(e => ({ tag:e.tagName.toLowerCase(), text:cut(e.innerText||e.value,26), vis:vis(e) }))
      .filter(x => x.text);
  const files = [...document.querySelectorAll('input[type=file]')]
      .map(e => ({ id:e.id||'', name:e.name||'', vis:vis(e), outer:cut(e.outerHTML,90) }));
  return {
    url: location.href,
    panelFound: !!panel,
    panelId: panel ? (panel.id || '') : '',
    panelDisplay: panel ? getComputedStyle(panel).display : '',
    panelText: panel ? cut(panel.innerText, 400) : '(パネル未検出)',
    fileInputs: files,
    clickables: clickables.slice(0, 25),
    totalButtons: document.querySelectorAll('button').length
  };
}
"""


_ATTACH_EXTS = (".docx", ".doc", ".docm", ".xlsx", ".xls", ".xlsm", ".pdf")


def _pdf_dir(out_folder, mapping):
    return os.path.join(out_folder, mapping.get("pdf_dir_name", "PDF変換"))


def _related_folder(out_folder, mapping, want_prefix):
    """案件フォルダ(例 2種関節系PRP_<ヒアリング名>)と同じ<ヒアリング名>を持つ別フォルダ
       (例 SOP_<ヒアリング名>)を返す。無ければ want_prefix で始まる最新フォルダ。"""
    parent = os.path.dirname(out_folder)
    base = os.path.basename(out_folder)
    case_pref = mapping.get("output_folder_contains", "")
    suffix = base[len(case_pref) + 1:] if (case_pref and base.startswith(case_pref + "_")) else ""
    if suffix:
        cand = os.path.join(parent, "%s_%s" % (want_prefix, suffix))
        if os.path.isdir(cand):
            return cand
    cands = [d for d in glob.glob(os.path.join(parent, want_prefix + "*")) if os.path.isdir(d)]
    return max(cands, key=os.path.getmtime) if cands else ""


def _find_source_files(folder, patterns):
    """folder 内で patterns(文字列/リスト)のいずれかを名前に含む対象ファイル一覧（~$除外）。"""
    if isinstance(patterns, str):
        patterns = [patterns]
    out = []
    for f in sorted(glob.glob(os.path.join(folder, "*"))):
        b = os.path.basename(f)
        if b.startswith("~$"):
            continue
        if os.path.splitext(b)[1].lower() not in _ATTACH_EXTS:
            continue
        if any(p and p in b for p in patterns):
            out.append(f)
    return out


def _ensure_pdf(src, pdf_dir, gray=True):
    """src(Word/Excel) → pdf_dir/<name>.pdf（白黒）。PDFは元より新しければ再変換しない(キャッシュ)。
       src が既にPDFならそのまま使う。戻り値: (pdfパス, 方式 or 'cached'/'error', エラー文)。"""
    if src.lower().endswith(".pdf"):
        return src, "pdf(そのまま)", ""
    os.makedirs(pdf_dir, exist_ok=True)
    dst = os.path.join(pdf_dir, os.path.splitext(os.path.basename(src))[0] + ".pdf")
    try:
        if os.path.exists(dst) and os.path.getmtime(dst) >= os.path.getmtime(src):
            return dst, "cached", ""
    except OSError:
        pass
    try:
        import to_pdf as T                              # 遅延import（未導入でも他機能を止めない）
        _, method = T.convert(src, dst, gray=gray)
        return dst, method, ""
    except Exception as e:
        return "", "error", repr(e)


def _resolve_attachments(mapping, out_folder, only=""):
    """attachments 設定 → [(slot, [ソースファイル…]), …]。from=SOP は別フォルダから拾う。
       only(スロット番号)指定でそのスロットのみ。"""
    sop = _related_folder(out_folder, mapping, mapping.get("sop_folder_contains", "SOP"))
    result = []
    for a in mapping.get("attachments", []):
        if not isinstance(a, dict):
            continue
        slot = str(a.get("slot", "")).strip()
        if not slot or (only and slot != only):
            continue
        folder = sop if a.get("from") == "SOP" else out_folder
        srcs = _find_source_files(folder, a.get("match", "")) if folder else []
        result.append((slot, srcs, a.get("from", "case")))
    return result


def _upload_one(page, sel, paths):
    """1スロットにファイル（複数可）をセットして「アップロード」を押す。
       戻り値: (ok:bool, msg:str)。※送信・申請は行わない（添付のみ）。"""
    if isinstance(paths, str):
        paths = [paths]
    fi = page.locator(sel)
    if fi.count() == 0:
        return False, "ファイル入力欄が見つかりません（%s）" % sel
    try:
        fi.first.set_input_files([os.path.abspath(p) for p in paths])
    except Exception as e:
        return False, "セット失敗 %r" % e
    page.wait_for_timeout(700)
    up = page.locator("xpath=//*[self::button or self::a]"
                      "[contains(normalize-space(.),'アップロード')]")
    for i in range(up.count()):
        try:
            if up.nth(i).is_visible():
                up.nth(i).click()
                page.wait_for_timeout(2200)
                return True, "アップロード済"
        except Exception:
            pass
    return True, "セットのみ（アップロードボタンが見つからず→手動で押してください）"


def _latest_receipt():
    """03_logs の最新レポートから 受付番号/パスワード を拾う（添付テストのショートカット用）。
       戻り値: (受付番号, パスワード, 元ファイル名)。見つからなければ空。"""
    files = sorted(glob.glob(os.path.join(LOGDIR, "web_autofill_report_*.txt")),
                   key=os.path.getmtime, reverse=True)
    for f in files:
        try:
            t = io.open(f, encoding="utf-8", errors="replace").read()
        except Exception:
            continue
        m1 = _re.search(r"受付番号:\s*([0-9A-Za-z\-]+)", t)
        if m1:
            m2 = _re.search(r"パスワード:\s*([0-9A-Za-z\-]+)", t)
            return m1.group(1), (m2.group(1) if m2 else ""), os.path.basename(f)
    return "", "", ""


def _prepare_slot_pdfs(mapping, out_folder, only="", R=None):
    """attachments 設定を解決し、各ソースを白黒PDF化して { slot: [pdf…] } を返す。
       R が渡されればレポート行を追記。PDF化はキャッシュ利用（元より新しければ再変換しない）。"""
    if R is None:
        R = []
    try:
        import to_pdf as T
        ready, msg = T.check_ready()
    except Exception as e:
        ready, msg = False, "to_pdf 読込失敗: %r" % e
    if not ready:
        R.append("  ★ %s" % msg)
        print("  ★ %s" % msg)
        return {}
    gray = mapping.get("pdf_grayscale", True)
    pdf_dir = _pdf_dir(out_folder, mapping)
    R.append("  PDF変換先: %s（白黒=%s）" % (pdf_dir, "はい" if gray else "いいえ"))
    resolved = _resolve_attachments(mapping, out_folder, only)
    slot_pdfs = {}
    tried = failed = 0
    for slot, srcs, frm in resolved:
        if not srcs:
            R.append("  スロット%-3s: 対象ファイル無し（from=%s）" % (slot, frm))
            continue
        for src in srcs:
            b = os.path.basename(src)
            tried += 1
            pdf, method, err = _ensure_pdf(src, pdf_dir, gray)
            if not pdf:
                failed += 1
                R.append("  ✗ 書類%-3s PDF化失敗: %s（%s）" % (slot, b, err))
                continue
            slot_pdfs.setdefault(slot, []).append(pdf)
            msg = "  ・書類%-3s ← %s → %s（%s）" % (slot, b, os.path.basename(pdf), method)
            R.append(msg)
            print(msg)
    if tried and failed == tried:                       # 全滅＝環境の問題を明示（クライアント初回対策）
        R.append("  ★ PDF化が全て失敗しました。次をご確認ください：")
        R.append("     ・Word/Excel（MS Office）がこのPCにインストールされているか")
        R.append("     ・『初回準備／Setup』を実行して部品(pywin32・PyMuPDF)を導入したか")
        R.append("     ・変換対象のWord/Excelを開いたままにしていないか（閉じてから再実行）")
    return slot_pdfs


def _upload_slots(page, slot_pdfs, R):
    """{slot:[pdf…]} を各 #btn_upload_fileupload<slot> に添付してアップロード。成功数を返す。"""
    okn = 0
    for slot in sorted(slot_pdfs, key=lambda x: int(x) if x.isdigit() else 999):
        pdfs = slot_pdfs[slot]
        ok, msg = _upload_one(page, "#btn_upload_fileupload%s" % slot, pdfs)
        R.append("  %s 書類%-3s ← %s  … %s"
                 % ("○" if ok else "×", slot,
                    " / ".join(os.path.basename(p) for p in pdfs), msg))
        if ok:
            okn += 1
    return okn


def run_pdf_convert(mapping):
    """★--pdf：添付対象の Word/Excel を白黒PDF化して PDF変換フォルダへ出すだけ（ブラウザ不要）。"""
    _ws, out_folder = load_output_ctx(mapping)
    R = ["=== PDF変換（--pdf）  (%s) ===" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M")]
    if not out_folder:
        R.append("アウトプット未検出。先に転記実行してください。")
        print("\n".join(R))
        return
    R.append("参照アウトプット: %s" % os.path.basename(out_folder))
    R.append("Ghostscript: %s" % (find_gs_note()))
    slot_pdfs = _prepare_slot_pdfs(mapping, out_folder, "", R)
    total = sum(len(v) for v in slot_pdfs.values())
    R.append("→ %dスロット・%dファイルをPDF化しました" % (len(slot_pdfs), total))
    print("\n".join(R))
    _write_report(R)


def find_gs_note():
    try:
        import to_pdf as T
        gs = T.find_ghostscript()
        return ("あり（高品質・ベクター白黒）: %s" % gs) if gs else "無し（PyMuPDFで画像白黒）"
    except Exception:
        return "判定不可"


def run_attach_one(page, mapping, only=""):
    """★添付（--attach）。一時保存済みの下書きを開いた状態で実行する。
       attachments 設定に従い、ソースを白黒PDF化して #btn_upload_fileupload<slot> へ貼る。
       only（例 "2"）でそのスロットだけ。※送信・申請はしない。"""
    _out_ws, out_folder = load_output_ctx(mapping)
    R = ["=== 添付（--attach）  (%s) ===" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M")]
    if not out_folder:
        R.append("アウトプット未検出。先に転記実行してください。")
        print("\n".join(R))
        return
    R.append("参照アウトプット: %s" % os.path.basename(out_folder))
    R.append("Ghostscript: %s" % find_gs_note())

    R.append("\n--- 白黒PDF化 ---")
    slot_pdfs = _prepare_slot_pdfs(mapping, out_folder, only, R)
    if not slot_pdfs:
        R.append("  添付できるPDFがありません")
        print("\n".join(R))
        _write_report(R)
        return

    _click_tab(page, mapping.get("attach_tab", "添付書類"))
    page.wait_for_timeout(2000)
    if page.locator("input[type='file']").count() == 0:
        try:
            d = page.evaluate(ATTACH_DIAG_JS)
        except Exception as e:
            d = {"error": repr(e)}
        R.append("\n★ ファイル入力欄がありません＝【一時保存した下書き】を開けていません。")
        R.append("  url: %s / panelFound=%s" % (d.get("url"), d.get("panelFound")))
        R.append("  → 受付番号・パスワードでログインし、添付書類タブを表示してから実行してください。")
        print("\n".join(R))
        _write_report(R)
        return

    R.append("\n--- 添付を実行（%dスロット）---" % len(slot_pdfs))
    okn = _upload_slots(page, slot_pdfs, R)
    R.append("→ %d/%d スロットを処理しました" % (okn, len(slot_pdfs)))
    R.append("※ 送信・申請はしていません。内容は画面でご確認ください。")

    try:
        shot = os.path.join(LOGDIR, "web_attach_%s.png" % datetime.datetime.now().strftime("%Y%m%d_%H%M"))
        page.screenshot(path=shot, full_page=True)
        R.append("\nスクリーンショット: %s" % shot)
    except Exception:
        pass
    path, _ = _write_report(R)
    print("\n".join(R))
    if path:
        print("\n診断を保存しました: %s" % path)


def _attach_files(page, mapping, out_folder):
    """④添付書類（--auto の一時保存後フローから呼ばれる）。
       attachments 設定に従い、ソース(Word/Excel)を白黒PDF化して #btn_upload_fileupload<slot> に添付。
       ※ 添付欄は【一時保存した下書き】でしか描画されないサイト仕様のため、必ず保存後に呼ぶこと。
       ※ 送信・申請はしない。"""
    lines = []
    if not out_folder:
        return lines
    lines.append("\n--- ④添付書類（白黒PDF化して添付）---")
    lines.append("  Ghostscript: %s" % find_gs_note())

    # 1) ソースを白黒PDF化（ブラウザ操作の前に済ませる：Office変換は時間がかかるため）
    slot_pdfs = _prepare_slot_pdfs(mapping, out_folder, "", lines)
    if not slot_pdfs:
        lines.append("  添付できるPDFがありません")
        return lines

    # 2) 添付書類タブを開いて貼る
    if not _click_tab(page, mapping.get("attach_tab", "添付書類")):
        lines.append("  添付書類タブを開けませんでした")
        return lines
    page.wait_for_timeout(1500)
    if page.locator("input[type='file']").count() == 0:
        lines.append("  ファイル入力欄がありません（一時保存済みの下書きを開けていない可能性）。")
        return lines
    okn = _upload_slots(page, slot_pdfs, lines)
    lines.append("  → %d/%d スロットを添付しました" % (okn, len(slot_pdfs)))
    return lines


# その欄がどのタブ（申請者情報/項目1〜7/添付書類）にあるかを返すJS＝エラー箇所の案内用
_TAB_OF_JS = r"""
(sel) => {
  let e = null;
  try { e = document.querySelector(sel); } catch (err) { return ''; }
  if (!e) return '';
  let p = e.closest('[role=tabpanel]');
  if (!p) { let n = e; while (n && n.parentElement) { n = n.parentElement;
              if (n.id && /^tab\d+$/.test(n.id)) { p = n; break; } } }
  if (!p) return '';
  const t = document.querySelector('[role=tab][aria-controls="' + p.id + '"]');
  return t ? (t.innerText || '').trim() : p.id;
}
"""


def _field_tab(page, sel):
    """欄のタブ名（例「項目3」）を返す。取れなければ空。"""
    if not sel:
        return ""
    try:
        return page.evaluate(_TAB_OF_JS, sel) or ""
    except Exception:
        return ""


def _write_output_summary(out_folder, receipt, plan, att_lines):
    """★02_output 直下に結果サマリを出力する（案件ごとの結果が一覧で並ぶように）。
       ・一時保存の受付番号／パスワード（再編集に必須）
       ・要対応（文字数超過・未入力・欄なし・未選択・エラー）を、項目名とタブ位置つきで列挙
       ・添付結果
       ※ パスは 02_output を前方一致で解決（【】付きフォルダ名でもOK／環境非依存）。
       戻り値: 出力先パス（失敗時は空）。"""
    if not out_folder:
        return ""
    def _sel(i):
        return i["fld"].get("selector") or i["fld"].get("label") or ""
    def _where(i):
        t = i.get("tab") or ""
        return ("%s タブ" % t) if t else "（タブ不明）"

    over = [i for i in plan if i.get("over")]
    empty = [i for i in plan if i["status"] == "empty"]
    notfound = [i for i in plan if i["status"] == "pending"]
    nochoice = [i for i in plan if i["status"] == "no-choice"]
    errored = [i for i in plan if i["status"] == "error"]
    placed = [i for i in plan if i["status"] == "placed"]

    L = []
    L.append("=" * 64)
    L.append(" Web転記 結果サマリ")
    L.append(" 実行: %s" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    L.append(" 案件: %s" % os.path.basename(out_folder))
    L.append("=" * 64)
    L.append("")
    L.append("■ 一時保存の受付情報（再編集・状況確認に必要。大切に保管）")
    if receipt.get("受付番号"):
        L.append("    受付番号  : %s" % receipt.get("受付番号"))
        L.append("    パスワード: %s" % receipt.get("パスワード", "(取得できず)"))
        L.append("    ログイン  : https://saiseiiryo.mhlw.go.jp/application/login/plan")
    else:
        L.append("    ※ 取得できませんでした（一時保存が完了していない可能性）")
    L.append("")
    L.append("■ 入力結果: %d / %d 欄" % (len(placed), len(plan)))
    L.append("")

    ng = over or empty or notfound or nochoice or errored
    if not ng:
        L.append("■ 要対応: なし（すべて転記できました）")
    else:
        L.append("■ 要対応（下記は画面で手直しが必要です）")
    if over:
        L.append("")
        L.append("  ―― 文字数超過：本文が入っていません（「文字数超過」と記入済み）――")
        for i in over:
            L.append("   ・%s" % i["desc"])
            L.append("       場所  : %s" % _where(i))
            L.append("       文字数: %d字 / 上限%d字（%d字超過）" % (i["len"], i["limit"], i["over"]))
            L.append("       対処  : アウトプットの本文を%d字以内に整えて再実行してください" % i["limit"])
    if empty:
        L.append("")
        L.append("  ―― 未入力：アウトプットに値がありません ――")
        for i in empty:
            L.append("   ・%-30s %s" % (i["desc"], _where(i)))
    if notfound:
        L.append("")
        L.append("  ―― 欄が見つからない：画面の構成が変わった可能性 ――")
        for i in notfound:
            L.append("   ・%-30s (selector=%s)" % (i["desc"], _sel(i)))
    if nochoice:
        L.append("")
        L.append("  ―― 選択肢が一致しない ――")
        for i in nochoice:
            L.append("   ・%-30s %s" % (i["desc"], _where(i)))
    if errored:
        L.append("")
        L.append("  ―― 入力時エラー ――")
        for i in errored:
            L.append("   ・%-30s %s  %s" % (i["desc"], _where(i), i.get("error", "")))
    if att_lines:
        L.append("")
        L.append("■ 添付書類")
        for x in att_lines:
            t = x.strip()
            if t and not t.startswith("---"):
                L.append("   %s" % t)
    L.append("")
    L.append("※ 送信・申請は行っていません。最終確認と送信は必ず人が行ってください。")

    # 02_output 直下に出力（案件名をファイル名に含めて一覧で見分けられるようにする）
    root = TX.resolve_dir("02_output", create=True)
    case = os.path.basename(out_folder)
    path = os.path.join(root, "Web転記結果_%s_%s.txt"
                        % (case, datetime.datetime.now().strftime("%Y%m%d_%H%M")))
    try:
        with io.open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(L))
        return path
    except Exception:
        return ""


def _write_report(lines):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(LOGDIR, "web_autofill_report_%s.txt" % ts)
    try:
        with io.open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    except Exception:
        path = ""
    return path, ts


def run_auto(page, mapping, hearing, kits):
    """全タブを自動で送りながら全欄を入力し、最後に一時保存。不備は箇所を報告。"""
    out_ws, out_folder = load_output_ctx(mapping)
    src_line = ("参照元(アウトプット): %s ／ 様式xlsx=%s"
                % (os.path.basename(out_folder), "あり" if out_ws is not None else "なし")) \
        if out_folder else "※ アウトプット未検出 → ヒアリングから取得（先に転記実行.batを推奨）"
    print(src_line)

    # ★1操作あたりの待機上限（既定30秒だと、別タブの要素を触った時に長時間フリーズするため）
    try:
        page.set_default_timeout(int(mapping.get("action_timeout_ms", 8000)))
    except Exception:
        pass

    plan = build_plan(mapping, hearing, kits, out_ws, out_folder)
    max_sections = int(mapping.get("max_sections", 25))
    print("自動一括入力を開始します（%d欄・最大%dタブ）…" % (len(plan), max_sections))

    def _fill_pass(label):
        placed_here = 0
        for item in plan:
            if item["status"] == "placed":
                continue
            st = _place_field(page, item)
            if st == "absent":
                continue                               # このタブには無い→別タブで再挑戦
            item["status"] = st                        # placed/empty/no-choice/error は確定
            if st == "placed":
                placed_here += 1
        print("  %s: %d欄入力" % (label, placed_here))

    tabs = mapping.get("tabs")
    if tabs:
        # ★タブ式フォーム：各タブを順に開いて処理（表示されて初めてチェック等が押せる）
        _fill_pass("現タブ")
        for tabname in tabs:
            if _click_tab(page, tabname):
                page.wait_for_timeout(500)
                _fill_pass(tabname)
            else:
                print("  （タブ「%s」が見つかりません）" % tabname)
        # 取りこぼし対策にもう一巡（前タブで未表示だったチェック等を回収）
        for tabname in tabs:
            if any(i["status"] not in ("placed", "empty", "no-choice", "error") for i in plan):
                if _click_tab(page, tabname):
                    page.wait_for_timeout(300)
                    _fill_pass(tabname + "(再)")

        # ★最終確定パス：一度入ったのにサイト側で解除されたチェックを入れ直す
        #   （他項目の入力やタブ切替で連動リセットされる事象への対策。最大2巡）
        _ON_JS = "(s) => { let o=false; document.querySelectorAll(s).forEach(e => { if (e.checked) o = true; }); return o; }"
        for _round in range(1):                          # 1巡のみ（何度も押すと戻るだけなので）
            reverted = []
            for item in plan:
                fld = item["fld"]
                ft = (fld.get("type") or "").lower()
                sel = _qualify_sel(fld.get("selector"), ft)      # 同名hiddenを除外して実体を見る
                if ft not in ("checkbox", "radio") or not sel:
                    continue
                want_on = True if ft == "radio" else \
                    (str(item["val"]) not in ("", "0", "false", "無", "×"))
                if not want_on:
                    continue
                try:
                    if not page.evaluate(_ON_JS, sel):
                        item["status"] = "pending"       # 解除されている→もう一度入れ直す
                        reverted.append(item["desc"])
                except Exception:
                    continue
            if not reverted:
                break
            print("  ↻ 解除されたチェックを入れ直します: %s" % ", ".join(reverted))
            for tabname in tabs:
                if any(i["status"] == "pending" for i in plan):
                    if _click_tab(page, tabname):
                        page.wait_for_timeout(400)
                        _fill_pass(tabname + "(確定)")
    else:
        for sec in range(max_sections):
            _fill_pass("タブ%d" % (sec + 1))
            if not _click_next(page, mapping):
                print("  （これ以上「次へ」が無いため、全タブ走査を終了）")
                break

    # ---- レポート集計 ----
    placed = [i for i in plan if i["status"] == "placed"]
    empty = [i for i in plan if i["status"] == "empty"]
    notfound = [i for i in plan if i["status"] == "pending"]     # どのタブにも無かった
    nochoice = [i for i in plan if i["status"] == "no-choice"]
    errored = [i for i in plan if i["status"] == "error"]

    # ★問題のある欄について「どのタブか」を今のうちに控える
    #   （一時保存すると画面が確認ページへ遷移し、以降は要素を辿れなくなるため）
    for i in plan:
        if i.get("over") or i["status"] in ("empty", "no-choice", "error"):
            i["tab"] = _field_tab(page, (i["fld"].get("selector") or "").strip())

    R = []
    R.append("=== Web自動一括入力 レポート  (%s) ===" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    R.append(src_line)
    R.append("入力できた欄: %d / %d" % (len(placed), len(plan)))

    def sel_of(i):
        return i["fld"].get("selector") or i["fld"].get("label") or ""

    if empty:
        R.append("\n【未入力：欄はあるが値が空】← データ元（アウトプット/ヒアリング）を確認")
        for i in empty:
            R.append("  - %s  (selector=%s)" % (i["desc"], sel_of(i)))
    if notfound:
        R.append("\n【見つからない：どのタブにも欄が無い】← selector/label の見直しが必要")
        for i in notfound:
            R.append("  - %s  (selector=%s)" % (i["desc"], sel_of(i)))
        # チェック/ラジオが absent のままの場合、なぜ可視ラベルが取れないのかを実DOMで診断
        _CHK_DIAG_JS = r"""
        (sel) => {
          const e = document.querySelector(sel);
          if (!e) return {found:false};
          const lab = e.closest('label');
          const rect = lab ? lab.getBoundingClientRect() : null;
          const cs = lab ? getComputedStyle(lab) : null;
          let panel = e.closest('[role=tabpanel]') || e.closest('div[id^=tab]');
          const pcs = panel ? getComputedStyle(panel) : null;
          return { found:true, checked:e.checked, disabled:e.disabled,
                   hasLabel: !!lab,
                   labW: rect ? Math.round(rect.width) : -1,
                   labH: rect ? Math.round(rect.height) : -1,
                   labDisplay: cs ? cs.display : '', labVis: cs ? cs.visibility : '',
                   panelId: panel ? (panel.id || panel.className || '') : '(none)',
                   panelDisplay: pcs ? pcs.display : '' };
        }"""
        for i in notfound:
            if (i["fld"].get("type") or "").lower() not in ("checkbox", "radio"):
                continue
            try:
                d = page.evaluate(_CHK_DIAG_JS, sel_of(i))
                R.append("    診断[%s]: %s" % (i["desc"], d))
            except Exception as e:
                R.append("    診断[%s]: 取得失敗 %r" % (i["desc"], e))
    if nochoice:
        R.append("\n【ラジオ未選択：選択肢が一致せず】← choice の文言を確認")
        for i in nochoice:
            R.append("  - %s  (choice=%s)" % (i["desc"], i["fld"].get("choice") or i["val"]))
    if errored:
        R.append("\n【入力時エラー】")
        for i in errored:
            R.append("  - %s  %s" % (i["desc"], i["error"]))

    # ---- チェック/ラジオの実状態を検証（本当に選択されたか）★JSで el.checked を直接読む（確実）----
    chk_targets = [(i, _qualify_sel(i["fld"].get("selector"), (i["fld"].get("type") or "").lower()))
                   for i in plan
                   if (i["fld"].get("type") or "").lower() in ("checkbox", "radio")
                   and (i["fld"].get("selector") or "").strip()]
    states = []
    try:
        states = page.evaluate(
            "(sels) => sels.map(s => { let els=[]; try{els=document.querySelectorAll(s);}catch(e){return {err:1};}"
            " let found=els.length>0, on=false; els.forEach(e=>{ if(e.checked) on=true; });"
            " return {found, on}; })",
            [s for _, s in chk_targets])
    except Exception:
        states = []
    chk_lines = []
    for (i, sel), st in zip(chk_targets, states or []):
        fld = i["fld"]
        ft = (fld.get("type") or "").lower()
        if not isinstance(st, dict) or st.get("err"):
            mark = "?確認不可"
        elif not st.get("found"):
            mark = "×欄なし"
        else:
            mark = "☑選択OK" if st.get("on") else "☐未選択"
        if ft == "checkbox":
            chk_lines.append("  %s  %s" % (mark, i["desc"]))
        else:
            chk_lines.append("  %s  %s → %s" % (mark, i["desc"], fld.get("choice") or i["val"]))
    if chk_lines:
        R.append("\n【チェック/ラジオ 実状態】← ☐未選択があれば画面で手直し")
        R.extend(chk_lines)

    # ---- ②文字数オーバー／注意 ----
    over = [i for i in plan if i.get("over")]
    trunc = [i for i in plan if i.get("truncated")]
    warns = [i for i in plan if i.get("warn_len") and not i.get("over")]
    if over:
        R.append("\n★【文字数オーバー】← ここは本文が入っていません。手直しが必要です")
        R.append("  （Webの数え方に合わせ、改行は2文字として計上しています）")
        for i in over:
            if i.get("marked_over"):
                tag = "→本文は入れず「%s」と記入（保存を通すため）" % i["val"]
            elif i.get("truncated"):
                tag = "→自動短縮で%d字に収めました" % i["limit"]
            else:
                tag = "→そのまま入力（保存が止まる恐れあり）"
            R.append("  - %s  %d字 / 上限%d字（%d字超過）\n      %s"
                     % (i["desc"], i["len"], i["limit"], i["over"], tag))
        R.append("  → 該当欄は、アウトプット側で本文を%d字以内に整えてから再実行してください。"
                 % (over[0]["limit"] if over else 4000))
    if warns:
        R.append("\n【文字数 注意】← 上限手前（改行が増えると溢れる恐れ）")
        for i in warns:
            R.append("  - %s  %d字 / 上限%d字" % (i["desc"], i["len"], i["limit"]))
    if trunc:
        R.append("※ 自動短縮した欄は、削られた内容が無いか必ず原文と照合してください。")

    # ---- 一時保存 ----（★添付はこの保存後に実施：サイト仕様「一時保存後に添付可能」）
    R.append("\n--- 一時保存（下書き保存）---")
    url_before = page.url
    clicked, btn_text = _click_save(page, mapping)
    if not clicked:
        R.append("一時保存ボタンが見つかりませんでした（送信系は自動的に除外しています）。")
        R.append("→ 画面の一時保存ボタンの selector を web_mapping.json の \"save_selector\" に設定してください。")
    else:
        R.append("一時保存ボタン「%s」をクリックしました。" % btn_text)

    site_errs = _collect_site_errors(page, mapping)
    if site_errs:
        R.append("\n★ サイトの検証メッセージ（不備）で止まっています：")
        for m in site_errs[:60]:
            R.append("  ● %s" % m)
        R.append("→ 上記の不備を修正のうえ、再実行するか画面で手直ししてください。")
    elif clicked:
        moved = (page.url != url_before)
        R.append("サイトの検証エラーは検出されませんでした（%s）。"
                 % ("保存後に画面遷移あり" if moved else "画面遷移なし"))
        R.append("→ 一時保存された可能性が高いですが、画面表示を必ずご確認ください。")

    # ---- 受付番号・パスワードの抽出（保存確認ページから）----
    receipt = {}
    att_lines = []
    if clicked:
        page.wait_for_timeout(800)
        receipt = _extract_receipt(page)
        if receipt:
            R.append("\n★ 一時保存の受付情報（再編集・状況確認に必要。大切に保管）:")
            for k in ("受付番号", "パスワード"):
                if receipt.get(k):
                    R.append("   %s: %s" % (k, receipt[k]))
            print("★ 受付番号: %s / パスワード: %s"
                  % (receipt.get("受付番号", "?"), receipt.get("パスワード", "?")))

    # ---- ④添付書類（★一時保存後にのみ可能：編集に戻る→添付→再保存）----
    if clicked and mapping.get("attach_enabled"):
        R.append("\n--- ④添付書類（一時保存後フロー）---")
        if _click_return_to_edit(page):
            R.append("「保存データ編集に戻る」で下書きを開きました。")
            page.wait_for_timeout(1200)
            att_lines = _attach_files(page, mapping, out_folder)
            R.extend(att_lines)
            page.wait_for_timeout(500)
            clicked2, _ = _click_save(page, mapping)   # 添付を保存するため再度一時保存
            page.wait_for_timeout(1000)
            R.append("添付後、再度一時保存しました。" if clicked2
                     else "※ 添付後の一時保存ボタンが見つかりませんでした。手動で保存してください。")
        else:
            R.append("「保存データ編集に戻る」ボタンが見つかりませんでした。")
    elif not mapping.get("attach_enabled"):
        R.append("\n--- ④添付書類（手動）---")
        R.append("  ※ 一時保存後、画面の「保存データ編集に戻る」→「添付書類」タブで下記を添付してください。")
        R.append("    （自動添付は保存後の編集ページでアップロード行が自動描画されず不可のため手動運用）")
        R.extend(_manual_attach_table(out_folder))

    # ---- スクリーンショット＆レポート保存 ----
    shot = ""
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        shot = os.path.join(LOGDIR, "web_autofill_%s.png" % ts)
        page.screenshot(path=shot, full_page=True)
    except Exception:
        shot = ""
    if shot:
        R.append("\nスクリーンショット: %s" % shot)
    rpath, _ = _write_report(R)

    # ---- ★アウトプットフォルダへ結果サマリ（受付番号・パスワード＋要対応の項目と場所）----
    spath = ""
    try:
        spath = _write_output_summary(out_folder, receipt, plan, att_lines)
    except Exception as e:
        print("結果サマリの出力に失敗: %r" % e)
    if spath:
        R.append("\n★ 結果サマリ（アウトプットに出力）: %s" % spath)

    print("\n" + "\n".join(R))
    if rpath:
        print("\nレポートを保存しました: %s" % rpath)
    if spath:
        print("★ 結果サマリを出力しました: %s" % spath)
    print("\n※ 送信は行っていません。最終確認と送信は必ず人が行ってください。")


def main():
    mode_dump = "--dump" in sys.argv
    mode_auto = "--auto" in sys.argv
    mode_attach = "--attach" in sys.argv
    mode_pdf = "--pdf" in sys.argv

    if mode_pdf:                                        # ブラウザ不要：PDF化だけ実行して終了
        run_pdf_convert(load_mapping())
        return

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("Playwright が未導入です。次を実行してください:")
        print("  pip install playwright")
        print("  playwright install chromium")
        return

    mapping = load_mapping()
    url = mapping.get("url") or "https://saiseiiryo.mhlw.go.jp/"
    hearing_sheet = mapping.get("hearing_sheet", "ヒアリングシート（PRP）")

    hearing = kits = None
    if not mode_dump and not mode_attach:              # 添付テストはヒアリング不要（起動を速く）
        hp = find_hearing()
        hearing = TX.Hearing(hp, hearing_sheet)
        kits = hearing.prp_kits()
        print("ヒアリング:", os.path.basename(hp))

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(PROFILE, headless=False,
                                                    viewport={"width": 1280, "height": 900})
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        # ★添付テストは「一時保存済みの下書き」で行う必要がある（サイト仕様: 保存後でないと
        #   アップロード部品が描画されない）。plan01へ直行すると“まっさらなフォーム”に戻って
        #   しまうため、サイトのトップを開いて利用者が下書きへ進めるようにする。
        start = url
        if mode_attach:
            # ★添付は「一時保存した下書き」でしか欄が出ないサイト仕様。下書きURLへの直行は
            #   セッション次第で表示できないため、ログイン画面を開いて手動ログインしてもらう。
            #   受付番号/パスワードは 03_logs の最新レポートから拾って表示する（貼り付け用）。
            start = mapping.get("attach_login_url",
                                "https://saiseiiryo.mhlw.go.jp/application/login/plan")
            no, pw, src = _latest_receipt()
            for a in sys.argv:
                if a.startswith("--no="):
                    no = a.split("=", 1)[1].strip()
                    pw = ""
            print("  ▼ログイン画面を開きます: %s" % start)
            if no:
                print("  ―― 直近の下書き（コピペ用）%s" % (("  ※%s より" % src) if src else ""))
                print("      受付番号  : %s" % no)
                print("      パスワード: %s" % (pw or "(レポートに記録なし)"))
        try:
            page.goto(start, wait_until="domcontentloaded", timeout=60000)
        except Exception:
            pass
        print("\n▼ ブラウザが開きました（このサイトはログイン不要）。")
        if mode_dump:
            input("  対象フォームを表示したら、このウィンドウで Enter（入力欄を抽出します）…")
            dump_fields(page)
        elif mode_attach:
            only = ""
            for a in sys.argv:
                if a.startswith("--only="):
                    only = a.split("=", 1)[1].strip()
            print("\n  ▼添付（%s）" % ("書類%s だけ" % only if only else "アウトプットの全ファイル"))
            print("    ①上の受付番号/パスワードでログイン（手動）")
            print("    ②「添付書類」タブを開き『ファイル選択』が見える状態にする")
            input("  そこまで進めたら、このウィンドウで Enter…")
            run_attach_one(page, mapping, only)
            input("\n  結果を確認したら Enter でブラウザを閉じます…")
        elif mode_auto:
            input("  plan01フォームの先頭タブを表示したら、このウィンドウで Enter（自動一括入力→一時保存を開始）…")
            print("  ※ 全タブを自動で送りながら入力し、最後に一時保存まで行います（送信はしません）。")
            run_auto(page, mapping, hearing, kits)
            input("\n  結果を確認したら Enter でブラウザを閉じます…")
        else:
            input("  plan01フォームを表示したら、このウィンドウで Enter（自動入力を開始）…")
            print("  ※ 複数ページのフォームは、入力後に「次の項目へ」で次を表示→Enter で続けて入力できます。")
            fill_fields(page, mapping, hearing, kits)
        ctx.close()


if __name__ == "__main__":
    main()
