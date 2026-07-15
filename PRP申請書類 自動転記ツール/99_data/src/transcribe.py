# -*- coding: utf-8 -*-
"""
PRP申請書類 自動転記ツール（複数シート対応・JSON駆動）

処理の流れ:
  1) 99_data/mapping.json を読み込む（緑セル→転記元の対応表）。
  2) 99_data/format_changes.json があれば動的にマッピングを上書き（オフライン自己修復の土台）。
  3) 01_input/ の最新ヒアリングシートを読み込み、索引（ラベル+セクション+出現順）を構築。
  4) 99_data/templates/ の各テンプレ(2種/3種)を【全シートを保ったまま】開き、
     mapping.json に従って緑セルへ実値を書き込む。
  5) 02_output/ に別名保存（元テンプレ・原本は上書きしない）。
  6) 99_data/_logs/ に転記ログを出力（クライアントの目に触れない位置）。

※ openpyxl で load_workbook→save するため、テンプレ内の全シート・書式・画像は保持される。
   1シートしか出ない不具合は、特定シートだけ書き出す実装が原因。本実装は全シート保持。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, glob, json, datetime, re, io

# ---- パス解決 ----
#   スクリプトの場所: <ルート>/99_data/src/transcribe.py
SRC_DIR = os.path.dirname(os.path.abspath(__file__))        # 99_data/src
DATA = os.path.dirname(SRC_DIR)                             # 99_data
BASE = os.path.dirname(DATA)                                # ツールのルート


def resolve_dir(prefix, create=False):
    """先頭が prefix に一致するルート直下フォルダを返す（【説明】付きでも拾う）。"""
    exact = os.path.join(BASE, prefix)
    if os.path.isdir(exact):
        return exact
    try:
        for name in sorted(os.listdir(BASE)):
            full = os.path.join(BASE, name)
            if os.path.isdir(full) and name.startswith(prefix):
                return full
    except OSError:
        pass
    if create:
        os.makedirs(exact, exist_ok=True)
    return exact


DIR_INPUT = resolve_dir("01_input")
DIR_OUTPUT = resolve_dir("02_output", create=True)
DIR_LOGS = resolve_dir("03_logs", create=True)
DIR_TPL = os.path.join(DATA, "テンプレート")
MAPPING_FILE = os.path.join(DATA, "マッピング", "mapping.json")
FORMAT_CHANGES_FILE = os.path.join(DATA, "マッピング", "format_changes.json")

# ステータス
ST_DONE = "転記済み"
ST_EMPTY = "空欄のため未転記"
ST_CHECK = "確認対象"
ST_UNMAP = "未割当(要マッピング)"
ST_NOSRC = "入力元なし"


def clean(s):
    return "" if s is None else str(s).strip()


# 連番マーカー（複数キット①②③…/最大15）
MARKS = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"


def is_placeholder(v):
    """ヒアリング未入力のダミー（〇〇 等）か判定。"""
    v = clean(v)
    if not v:
        return True
    # 〇/○/●/＊/* だけ、または「〇〇」を含む雛形値
    return ("〇" in v) or ("○" in v) or bool(re.fullmatch(r"[●＊*\s]+", v))


def wareki(dt):
    y = dt.year - 2018
    yy = "元" if y == 1 else str(y)
    return "令和%s年%d月%d日" % (yy, dt.month, dt.day)


# =========================================================================
#  ヒアリングブックの索引 + クロスシート参照
# =========================================================================
class Hearing:
    def __init__(self, path, hearing_sheet):
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self.entries = []
        try:
            ws = self.wb[hearing_sheet]
        except KeyError:
            ws = None
        self.ws = ws
        if ws is not None:
            cur = ""
            for r in range(1, ws.max_row + 1):
                a = ws.cell(r, 1).value
                b = ws.cell(r, 2).value
                c = ws.cell(r, 3).value
                if isinstance(a, str) and a.strip():
                    cur = a.strip()
                if isinstance(b, str) and b.strip():
                    self.entries.append(dict(section=cur, label=b.strip(), value=clean(c)))

    # --- ラベル+セクション+出現順 ---
    def lookup(self, label, section="", occ=1):
        label = clean(label); section = clean(section)
        if not label:
            return None
        hits = [e for e in self.entries
                if label in e["label"] and (not section or section in e["section"])]
        if len(hits) >= occ >= 1:
            return hits[occ - 1]["value"]
        return None

    # --- 別シートのセル直接参照 ---
    def sheet_cell(self, sheet, cell):
        try:
            ws = self.wb[sheet]
        except KeyError:
            return None
        try:
            return clean(ws[cell].value)
        except Exception:
            return None

    # --- 委員会名 → 審査委員会シートで認定番号(B列) ---
    def committee_number(self, name):
        return self.committee_field(name, 2)

    # --- 委員会名 → 審査委員会シートの任意列（1名称/2認定番号/3連絡先名/4住所/5TEL/6メール）---
    def committee_field(self, name, col):
        name = clean(name)
        if not name:
            return ""
        try:
            ws = self.wb["審査委員会"]
        except KeyError:
            return ""
        for row in ws.iter_rows(values_only=True):
            if row and clean(row[0]) == name:
                return clean(row[col - 1]) if len(row) >= col else ""
        return ""

    # --- ヒアリングで選択されたPRPキット名（複数キット対応・雛形値は除外）---
    def prp_kits(self):
        kits = []
        for e in self.entries:
            if "PRPキットメーカー" in e["label"] and not is_placeholder(e["value"]):
                v = clean(e["value"])
                if v not in kits:
                    kits.append(v)
        return kits

    # --- PRPメーカーシートで キット名→採取(C)/加工(D)/投与(E) ---
    #     part: "採取"|"加工"|"投与"|"名称"|"採血量"
    _KIT_COL = {"名称": 1, "採血量": 2, "採取": 3, "加工": 4, "投与": 5}

    def kit_method(self, kit_name, part):
        name = clean(kit_name)
        if not name:
            return ""
        try:
            ws = self.wb["PRPメーカー"]
        except KeyError:
            return ""
        col = self._KIT_COL.get(part, 4)
        norm = lambda s: re.sub(r"\s+", "", clean(s))
        rows = list(ws.iter_rows(values_only=True))

        def pick(row):
            val = clean(row[col - 1]) if row and len(row) >= col else ""
            if part == "投与":           # 投与方法は見出し【…】を除去して本文だけ
                val = re.sub(r"^\s*【[^】]*】\s*", "", val)
            return val
        for row in rows:                     # 正規化一致
            if row and norm(row[0]) == norm(name):
                return pick(row)
        for row in rows:                     # 部分一致
            a = norm(row[0]) if row else ""
            if a and (a in norm(name) or norm(name) in a):
                return pick(row)
        return ""

    # --- 複数キットを①②③…で連番連結（カテゴリ非依存の汎用ブロック生成）---
    def kit_block(self, part, with_name=False, numbered=True, joiner="\n\n"):
        kits = self.prp_kits()
        blocks = []
        for i, k in enumerate(kits):
            body = self.kit_method(k, part)
            if not body and not with_name:
                continue
            mark = ""
            if numbered:
                mark = MARKS[i] if i < len(MARKS) else "(%d)" % (i + 1)
            if with_name:                    # 「①キット名 + 本文」
                head = mark + clean(self.kit_method(k, "名称") or k)
                blocks.append((head + ("\n" + body if body else "")).strip())
            else:                            # 「①本文」（先頭の箇条記号は連番に置換）
                b = re.sub(r"^\s*[■●◆・]\s*", "", body) if mark else body
                blocks.append((mark + b) if mark else b)
        # 全キットで内容が同一（投与方法など）の場合は重複させず1つに畳む
        uniq = []
        for b in blocks:
            core = re.sub(r"^[①-⑮]|^\(\d+\)", "", b).strip()
            if not any(re.sub(r"^[①-⑮]|^\(\d+\)", "", x).strip() == core for x in uniq):
                uniq.append(b)
        use = blocks if (with_name or len(uniq) == len(blocks)) else (
            [re.sub(r"^[①-⑮]|^\(\d+\)", "", uniq[0]).strip()] if len(uniq) == 1 else blocks)
        return joiner.join([b for b in use if b])

    # --- 医師略歴書 B7（医師免許＋学歴＋職歴の合体セル）を3分割 ---
    #     返り値: (医師免許, 学歴（大学）, 職歴)
    def rireki_b7_parts(self):
        b7 = clean(self.sheet_cell("略歴書（PRP）", "B7"))
        lic, edu, car = [], [], []
        for line in b7.split("\n"):
            s = line.strip()
            if not s:
                continue
            if ("医籍番号" in s) or ("免許取得日" in s):
                lic.append(s)
            elif ("大学" in s) and ("卒" in s):     # 卒業＝学歴（大学）。※職歴の「○○大学…病院」は卒無しで除外
                edu.append(s)
            else:
                car.append(s)
        return ("　".join(lic), "\n".join(edu), "\n".join(car))

    # --- 採血量（〇ml単体はそのまま／〇ml-〇ml等の範囲は「（使用キットにより異なる）」を付す）---
    def blood_volume(self):
        v = clean(self.lookup("必要な採血量", "", 1))
        if not v:
            return ""
        # 既に注記があれば二重付与しない
        if "使用キットにより異なる" in v:
            return v
        if re.search(r"[-~〜～−]", v):     # 範囲指定（ハイフン/波ダッシュ各種）
            return v + "（使用キットにより異なる）"
        return v

    # --- 治療価格ブロック（選択キット×対象再生医療の治療費を①②③で列挙）---
    #     saisei_idx: 1=再生医療①(2種) / 2=再生医療②(3種) …
    #     ヒアリングは「PRPキットメーカー①」ブロックごとに
    #     「再生医療①/②/③（右記タブから選択）の治療費の設定（税別）」を持つ。
    def kit_price_block(self, saisei_idx, joiner="\n"):
        # 各キットは「自分の出現番号」のマークで治療費を持つ：
        #   「再生医療[キット番号]（右記タブから選択）の治療費の設定（税別）」
        # ヒアリングのキット出現順にマークを対応させ、キット別の料金を列挙する。
        # （旧実装は saisei_idx 固定マーク＋occ でキット②以降を取りこぼしていた＝A3バグ）
        kit_vals = [clean(e["value"]) for e in self.entries
                    if "PRPキットメーカー" in e["label"] and not is_placeholder(e["value"])]
        lines, seen = [], []
        for orig_i, kit in enumerate(kit_vals):
            if not kit:
                continue
            name = re.sub(r"\s*（[^（）]*）\s*$", "", kit).strip()   # 末尾の（メーカー名）を除去
            if name in seen:                                        # 同一キットは1回だけ（連番展開と同じ畳み込み）
                continue
            mark_orig = MARKS[orig_i] if orig_i < len(MARKS) else "(%d)" % (orig_i + 1)
            price_label = "再生医療%s（右記タブから選択）の治療費の設定（税別）" % mark_orig
            price = clean(self.lookup(price_label, "", 1))          # マークが一意なので occ=1
            if not price or is_placeholder(price):
                continue
            mark_out = MARKS[len(seen)] if len(seen) < len(MARKS) else "(%d)" % (len(seen) + 1)
            seen.append(name)
            lines.append("%s%sを用いた治療：%s" % (mark_out, name, price))
        return joiner.join(lines)

    # --- 再生医療を行う医師の一覧（雛形値・重複を除外、最大20名）---
    #     「再生医療を行う医師…」で始まる人員欄のみ対象（履歴書案内行などは除外）
    def doctors(self):
        names = []
        for e in self.entries:
            if not e["label"].startswith("再生医療を行う医師"):
                continue
            if is_placeholder(e["value"]):
                continue
            v = clean(e["value"])
            if v and v not in names:
                names.append(v)
        return names[:20]


# =========================================================================
#  format_changes.json（オフライン自己修復の土台）
# =========================================================================
def apply_format_changes(mapping, run_log):
    """様式変更・項目名変更に追従するため、転記前にマッピングを動的上書きする。
       スケルトン実装。format_changes.json が無ければ何もしない。
       対応する変更タイプ:
         label_renames: ヒアリング項目名の改称（旧→新）
         cell_moves   : 出力先セルの移動（様式変更）
         disable      : 特定エントリの無効化
         set_source   : エントリのソースを差し替え
    """
    if not os.path.exists(FORMAT_CHANGES_FILE):
        return mapping
    try:
        with io.open(FORMAT_CHANGES_FILE, encoding="utf-8") as f:
            ch = json.load(f)
    except Exception as e:
        run_log.append("format_changes.json 読込失敗: %r" % e)
        return mapping

    renames = {clean(r.get("old")): r.get("new") for r in ch.get("label_renames", []) if r.get("old")}
    moves = ch.get("cell_moves", [])
    disables = ch.get("disable", [])
    setsrc = ch.get("set_source", [])
    n = 0

    for doc_key, doc in mapping.get("documents", {}).items():
        for e in doc.get("entries", []):
            src = e.get("source", {})
            # 1) ラベル改称
            if src.get("t") == "hearing" and clean(src.get("label")) in renames:
                old = clean(src.get("label")); src["label"] = renames[old]; n += 1
                run_log.append("[format_changes] %s: ラベル改称 '%s'→'%s'" % (doc_key, old, src["label"]))
            # 2) セル移動
            for mv in moves:
                if mv.get("document") == doc_key and mv.get("var") == e.get("var") and mv.get("new_cell"):
                    e["cell"] = mv["new_cell"]; n += 1
                    run_log.append("[format_changes] %s: %s のセルを %s へ移動" % (doc_key, e["var"], e["cell"]))
            # 3) ソース差し替え
            for ss in setsrc:
                if ss.get("document") == doc_key and ss.get("var") == e.get("var") and ss.get("source"):
                    e["source"] = ss["source"]; n += 1
                    run_log.append("[format_changes] %s: %s のソースを差し替え" % (doc_key, e["var"]))
            # 4) 無効化
            for ds in disables:
                if ds.get("document") == doc_key and (
                        ds.get("var") == e.get("var") or
                        (ds.get("sheet") == e.get("sheet") and ds.get("cell") == e.get("cell"))):
                    e["_disabled"] = True; n += 1
    if n:
        run_log.append("format_changes.json を適用（%d件）" % n)
    return mapping


# =========================================================================
#  ソース解決
# =========================================================================
def resolve(src, hearing, idx_kits):
    """ソース定義→(値, ステータス)"""
    t = src.get("t")
    if t == "hearing":
        v = hearing.lookup(src.get("label"), src.get("section", ""), int(src.get("occ", 1)))
        return (v, ST_DONE) if v not in (None, "") else ("", ST_EMPTY if v == "" else ST_NOSRC)
    if t == "sheet":
        v = hearing.sheet_cell(src.get("sheet"), src.get("cell"))
        return (v, ST_DONE) if v not in (None, "") else ("", ST_EMPTY)
    if t == "committee":
        cname = hearing.lookup("認定再生医療等委員会の名称", "審査委員会", 1)
        colmap = {"認定番号": 2, "連絡先名": 3, "住所": 4, "TEL": 5, "メール": 6}
        col = colmap.get(src.get("field", "認定番号"), 2)
        v = hearing.committee_field(cname, col)
        return (v, ST_DONE) if v else ("", ST_EMPTY)
    if t == "kit":
        part = src.get("part", "加工"); i = int(src.get("idx", 0))
        kits = idx_kits
        # idx=0 または mode=block → 全キットを①②③…で連番連結（汎用）
        if i == 0 or src.get("mode") == "block":
            v = hearing.kit_block(
                part,
                with_name=bool(src.get("with_name")),
                numbered=bool(src.get("numbered", True)),
                joiner=src.get("joiner", "\n\n"),
            )
            return (v, ST_CHECK) if v else ("", ST_EMPTY)
        # idx>=1 → そのキット単体（書類2の固定スロット用）
        if i <= len(kits):
            mark = (MARKS[i - 1] if i - 1 < len(MARKS) else "(%d)" % i) if src.get("with_mark") else ""
            if part == "名称":
                body = clean(hearing.kit_method(kits[i - 1], "名称") or kits[i - 1])
            else:
                body = hearing.kit_method(kits[i - 1], part)
            return ((mark + body) if body else "", ST_CHECK) if body else ("", ST_EMPTY)
        return ("", ST_EMPTY)   # 選択キット数より多いスロット → 空（行追加は体裁側）
    if t == "compose":
        # 複数項目を改行で連結（「名称・所在地・連絡先が一行」問題の分割整形）
        parts = []
        for ln in src.get("lines", []):
            sub = ln.get("source") or {
                "t": "hearing", "label": ln.get("label"),
                "section": ln.get("section", ""), "occ": ln.get("occ", 1)}
            v, _ = resolve(sub, hearing, idx_kits)
            v = clean(v)
            if v:
                parts.append(clean(ln.get("prefix", "")) + v + clean(ln.get("suffix", "")))
            elif ln.get("keep_empty"):
                parts.append(clean(ln.get("prefix", "")))
        return ("\n".join(parts), ST_CHECK) if parts else ("", ST_EMPTY)
    if t == "doctor":
        # 再生医療を行う医師（複数名）。field: name|affiliation|role
        i = int(src.get("idx", 1)); field = src.get("field", "name")
        docs = hearing.doctors()
        if i <= len(docs):
            if field == "name":
                return (docs[i - 1], ST_DONE)
            if field == "affiliation":
                v = hearing.lookup("医療機関/名称（診療所開設届上）", "法人/医療機関", 1)
                return (v, ST_CHECK) if v else ("", ST_EMPTY)
            if field == "role":
                return (clean(src.get("value", "院長")) if i == 1 else "", ST_CHECK)
        return ("", ST_EMPTY)
    if t == "date":
        return (wareki(datetime.datetime.now().date()), ST_CHECK)
    if t == "today":
        # 実行日（作業日）。fmt: slash="YYYY/M/D" / ymd="YYYY年M月D日" / wareki
        d = datetime.datetime.now()
        fmt = src.get("fmt", "ymd")
        if fmt == "slash":
            return ("%d/%d/%d" % (d.year, d.month, d.day), ST_CHECK)
        if fmt == "wareki":
            return (wareki(d.date()), ST_CHECK)
        return ("%d年%d月%d日" % (d.year, d.month, d.day), ST_CHECK)
    if t == "blood_volume":
        v = hearing.blood_volume()
        return (v, ST_DONE) if v else ("", ST_EMPTY)
    if t == "kitprice":
        v = hearing.kit_price_block(int(src.get("saisei", 1)), joiner=src.get("joiner", "\n"))
        return (v, ST_CHECK) if v else ("", ST_EMPTY)
    if t == "fixed":
        return (clean(src.get("value")), ST_CHECK)
    if t == "unknown":
        return (None, ST_UNMAP)
    return (None, ST_NOSRC)


# 医師略歴書シートの項目名 → セル対応（トークン名で参照）。
# 医師免許/学歴（大学）/職歴 は合体セルB7を分割して充てるため、ここには含めない。
_RIREKI_CELLS = {
    "氏名（ふりがな）": "B2", "氏名": "B3", "生年月日": "B4", "所属": "B5", "役職": "B6",
    "専門分野": "B8", "所属学会等": "B9", "認定医等の資格": "B10",
    "臨床経験及び研究に関する実績": "B11",
}


def resolve_token_by_name(name, hearing, kits, saisei):
    """トークン名（{{...}} の中身）から実値を解決する汎用リゾルバ。
       前世代の広範なトークン（キット製造方法・略歴書・委員会・作業日 等）も取りこぼさない。
       返り値は文字列（未解決/空は ''）。"""
    name = clean(name)
    if not name:
        return ""

    def kidx():
        m = re.search(r"メーカー([①-⑮])", name)
        if m:
            return MARKS.find(m.group(1))
        m = re.search(r"メーカー([A-O])", name)
        if m:
            return ord(m.group(1)) - ord("A")
        return -1

    # 作業日（日付）
    if "作業日" in name and ("スラッシュ" in name or "/" in name):
        return resolve({"t": "today", "fmt": "slash"}, hearing, kits)[0]
    if ("作業日" in name) or ("作成日" in name) or ("施行日" in name) or ("制定" in name):
        return resolve({"t": "today", "fmt": "ymd"}, hearing, kits)[0]
    # 採血量・治療価格
    if name == "必要な採血量":
        return hearing.blood_volume()
    if name == "治療価格":
        return hearing.kit_price_block(saisei)
    if "治療費の設定" in name:
        idx = 2 if "②" in name else (3 if "③" in name else 1)
        return clean(hearing.lookup("再生医療%s（右記タブから選択）の治療費の設定（税別）" % MARKS[idx - 1], "", 1)) or ""
    # キット製造方法（スロット番号付き＝単一キット／無印＝全キットのブロック）
    if ("採取" in name) and ("メーカー" in name or "キット" in name or "細胞" in name):
        i = kidx()
        return (hearing.kit_method(kits[i], "採取") if i < len(kits) else "") if i >= 0 \
            else hearing.kit_block("採取", numbered=True)
    if ("加工の方法" in name) or (name == "加工方法"):
        i = kidx()
        return (hearing.kit_method(kits[i], "加工") if i < len(kits) else "") if i >= 0 \
            else hearing.kit_block("加工", numbered=True)
    if ("投与の方法" in name) or (name == "投与方法"):
        i = kidx()
        return (hearing.kit_method(kits[i], "投与") if i < len(kits) else "") if i >= 0 \
            else hearing.kit_block("投与", numbered=True)
    # キット名（PRPキットメーカー①（右記タブ…）
    if "PRPキットメーカー" in name and "右記タブ" in name:
        ks = hearing.prp_kits(); i = kidx()
        return re.sub(r"\s*（[^（）]*）\s*$", "", ks[i]).strip() if 0 <= i < len(ks) else ""
    # 認定再生医療等委員会
    if "認定再生医療等委員会" in name:
        cname = hearing.lookup("認定再生医療等委員会の名称", "審査委員会", 1)
        if "認定番号" in name:
            return hearing.committee_field(cname, 2)
        if "連絡先名" in name:
            return hearing.committee_field(cname, 3)
        if "住所" in name:
            return hearing.committee_field(cname, 4)
        if "TEL" in name:
            return hearing.committee_field(cname, 5)
        if "メール" in name:
            return hearing.committee_field(cname, 6)
        return clean(cname)
    # 医師略歴書：医師免許/学歴（大学）/職歴 は合体セルB7を分割
    if name in ("医師免許", "学歴（大学）", "職歴"):
        lic, edu, car = hearing.rireki_b7_parts()
        return {"医師免許": lic, "学歴（大学）": edu, "職歴": car}[name]
    # 医師略歴書（その他はシート直接参照）
    if name in _RIREKI_CELLS:
        return clean(hearing.sheet_cell("略歴書（PRP）", _RIREKI_CELLS[name]))
    # 実施責任者の氏名
    if "実施責任者" in name and "氏名" in name:
        return clean(hearing.lookup("氏名", "実施責任者", 1)) or ""
    # 汎用：ヒアリング項目名として直接参照（表記ゆれ より→から を吸収）
    v = hearing.lookup(name.replace("より", "から"), "", 1)
    return clean(v) if v not in (None, "") else ""


def safe_set(ws, coord, val):
    """結合セルでも安全に値を書く。非アンカーの結合セルが指定されたら左上アンカーへ書く。"""
    from openpyxl.utils import range_boundaries, get_column_letter as _gl
    cell = ws[coord]
    if cell.__class__.__name__ == "MergedCell":
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                min_col, min_row, _, _ = range_boundaries(str(rng))
                ws.cell(row=min_row, column=min_col).value = val
                return True
        return False   # 書けない結合セル（まれ）はスキップ
    cell.value = val
    return True


def token_for(src):
    """source定義 → トークン名（ヒアリング項目名ベース）。None=トークン化しない。
       テンプレの {{トークン}} 表記と、転記エンジンの穴埋めキーを一致させるための唯一の定義。"""
    t = src.get("t")
    if t == "hearing":
        return (src.get("label") or "").strip()
    if t == "committee":
        f = src.get("field", "認定番号")
        return {
            "認定番号": "認定再生医療等委員会の認定番号",
            "連絡先名": "認定再生医療等委員会の連絡先名称",
            "住所": "認定再生医療等委員会の住所",
            "TEL": "認定再生医療等委員会の連絡先TEL",
            "メール": "認定再生医療等委員会の連絡先メールアドレス",
        }.get(f, "認定再生医療等委員会の認定番号")
    if t == "kit":
        return {"採取": "採取方法", "加工": "加工方法", "投与": "投与方法",
                "名称": "キット名称"}.get(src.get("part", "加工"), "加工方法")
    if t == "today":
        return "作業日" if src.get("fmt") == "ymd" else "作業日スラッシュ"
    if t == "compose":
        pre = (src.get("lines", [{}])[0].get("prefix", "")).replace("：", "").strip()
        return pre or "作成年月日"
    if t == "blood_volume":
        return "必要な採血量"
    if t == "kitprice":
        return "治療価格"
    if t in ("unknown", "sheet", "doctor"):
        return None
    return None


_TOKEN_RE = re.compile(r"\{\{[^{}]*\}\}")


def clear_tokens(wb):
    """テンプレに残った未充足 {{...}} プレースホルダを空欄化（部分一致も除去）。"""
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "{{" in c.value and "}}" in c.value:
                    new = _TOKEN_RE.sub("", c.value).strip()
                    if new != c.value:
                        safe_set(ws, c.coordinate, new if new else None)
                        n += 1
    return n


def apply_tf(val, tf):
    if val is None:
        return val
    val = clean(val)
    if tf == "postal":
        return ("〒" + val) if (val and not val.startswith("〒")) else val
    if isinstance(tf, str) and tf.startswith("prefix:"):
        pre = tf[len("prefix:"):]
        return (pre + val) if val else val
    return val


# =========================================================================
#  メイン
# =========================================================================
def find_hearing():
    if len(sys.argv) > 1 and clean(sys.argv[1]):
        return sys.argv[1]
    cands = [c for c in glob.glob(os.path.join(DIR_INPUT, "*.xlsx"))
             if not os.path.basename(c).startswith("~$")]
    # ヒアリング本体を優先（略歴書など他のExcelが 01_input に混在しても誤選択しない）
    named = [c for c in cands if "ヒアリング" in os.path.basename(c)]
    pool = named or [c for c in cands if "略歴" not in os.path.basename(c)] or cands
    pool.sort(key=os.path.getmtime, reverse=True)
    if pool:
        return pool[0]
    raise FileNotFoundError("01_input にヒアリングシート(.xlsx)を入れてください")


def main():
    run_dt = datetime.datetime.now()
    stamp = run_dt.strftime("%Y%m%d_%H%M%S")
    run_log = []

    # ヒアリングシートのタブ名（docs_config優先、無ければ既定）
    hearing_sheet = "ヒアリングシート（PRP）"
    cfgp = os.path.join(DATA, "マッピング", "docs_config.json")
    if os.path.exists(cfgp):
        try:
            hearing_sheet = json.load(io.open(cfgp, encoding="utf-8")).get("hearing_sheet", hearing_sheet)
        except Exception:
            pass

    hearing_path = find_hearing()
    hearing = Hearing(hearing_path, hearing_sheet)
    if hearing.ws is None:
        raise RuntimeError("ヒアリングシートのタブ '%s' が見つかりません" % hearing_sheet)

    # --- フォルダ型テンプレ（2種関節系PRP / 3種筋腱靭帯系PRP / SOP）へ転記 ---
    #     様式=Excelセル転記、その他=Wordの一括置換、PDF等=無変更で同梱
    out_paths = []
    log_rows = []
    try:
        from transcribe_docs import run_docs
        folders, log_rows = run_docs(hearing, hearing_path, DIR_TPL, DIR_OUTPUT, run_log, run_dt)
        out_paths = [(f, "フォルダ") for f in folders]
    except Exception as e:
        import traceback
        run_log.append("[docs] 転記でエラー: %r" % e)
        run_log.append(traceback.format_exc())

    log_path = write_log(log_rows, run_log, hearing_path, out_paths, run_dt, stamp)

    print("=== 転記完了（フォルダ型：Excel様式＋Word書類＋PDF同梱）===")
    print("入力 :", hearing_path)
    for p, ns in out_paths:
        print("出力 : %s （%s）" % (p, ns))
    print("ログ :", log_path)
    for line in run_log:
        print("  -", line)


def make_row(run_dt, doc_key, e, content, status, note):
    content = clean(content)
    return {
        "転記日時": run_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "文書": doc_key,
        "シート": clean(e.get("sheet")),
        "セル": clean(e.get("cell")),
        "変数": clean(e.get("var")),
        "転記内容": (content[:120] + "…") if len(content) > 120 else content,
        "処理結果": status,
        "確認": "要確認" if status in (ST_CHECK, ST_UNMAP, ST_NOSRC) else "OK",
        "備考": clean(note),
    }


def write_log(rows, run_log, hearing_path, out_paths, run_dt, stamp):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "転記ログ"
    cols = ["転記日時", "文書", "シート", "セル", "変数", "転記内容", "処理結果", "確認", "備考"]
    hf = PatternFill("solid", fgColor="4472C4"); hfont = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(cols, 1):
        c = ws.cell(1, j, h); c.fill = hf; c.font = hfont; c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center")
    cmap = {ST_DONE: "E2EFDA", ST_EMPTY: "FCE4D6", ST_CHECK: "FFF2CC", ST_UNMAP: "F8CBAD", ST_NOSRC: "FFF2CC"}
    for i, row in enumerate(rows, 2):
        for j, h in enumerate(cols, 1):
            c = ws.cell(i, j, row.get(h)); c.border = bd
            c.alignment = Alignment(vertical="center", wrap_text=True)
        fc = cmap.get(row["処理結果"])
        if fc:
            ws.cell(i, 7).fill = PatternFill("solid", fgColor=fc)
    for j, w in enumerate([19, 16, 26, 7, 18, 46, 16, 10, 32], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("実行情報")
    info = [("転記日時", run_dt.strftime("%Y-%m-%d %H:%M:%S")), ("入力", hearing_path)]
    for p, ns in out_paths:
        info.append(("出力(%s)" % ns, p))
    info.append(("メモ", " / ".join(run_log)))
    for i, (k, v) in enumerate(info, 1):
        ws2.cell(i, 1, k).font = Font(bold=True); ws2.cell(i, 2, v)
    ws2.column_dimensions["A"].width = 18; ws2.column_dimensions["B"].width = 90

    path = os.path.join(DIR_LOGS, "転記ログ_%s.xlsx" % stamp)
    wb.save(path)
    return path


if __name__ == "__main__":
    main()
