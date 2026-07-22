# -*- coding: utf-8 -*-
"""
PRP申請書類 自動転記ツール ― ローカルWebダッシュボード

既存の 99_data/src/transcribe.py には一切手を加えず、その実行を
ブラウザ画面から操作できるようにする薄いラッパー（Flask）。

  - ヒアリングシート(.xlsx)をドラッグ&ドロップ → 01_input へ保存
  - 「転記実行」→ transcribe.py をサブプロセス実行し、進捗をライブ表示(SSE)
  - 実行後、03_logs の最新ログExcelを解析して 統計・要確認(黄色)項目 を表示
  - 02_output のフォルダをZIPでダウンロード
  - 過去の実行ログ履歴を一覧・再表示

起動:  python3 98_dashboard/app.py   （ブラウザで http://127.0.0.1:8765 ）
"""
import os
import io
import sys
import glob
import json
import shutil
import zipfile
import datetime
import threading
import subprocess

from flask import (
    Flask, request, Response, jsonify,
    send_file, stream_with_context,
)
import openpyxl

# ---------------------------------------------------------------------------
# パス解決（transcribe.py と同じ流儀で、ツールのルートを基準にする）
#   このファイル: <ルート>/98_dashboard/app.py
# ---------------------------------------------------------------------------
DASH_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.dirname(DASH_DIR)                      # ツールのルート
DATA = os.path.join(BASE, "99_data")
TRANSCRIBE = os.path.join(DATA, "src", "transcribe.py")
WEB_FILL = os.path.join(DATA, "src", "web_fill.py")
WEB_MAPPING = os.path.join(DATA, "マッピング", "web_mapping.json")
DOCS_CONFIG = os.path.join(DATA, "マッピング", "docs_config.json")


def resolve_dir(prefix, create=False):
    """先頭が prefix に一致するルート直下フォルダを返す（【説明】付きでも拾う）。"""
    exact = os.path.join(BASE, prefix)
    if os.path.isdir(exact):
        return exact
    try:
        for name in sorted(os.listdir(BASE)):
            full = os.path.join(BASE, name)
            if os.path.isdir(full) and name.startswith(prefix):
                return full
    except OSError:
        pass
    if create:
        os.makedirs(exact, exist_ok=True)
    return exact


DIR_INPUT = resolve_dir("01_input", create=True)
DIR_OUTPUT = resolve_dir("02_output", create=True)
DIR_LOGS = resolve_dir("03_logs", create=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024   # 100MB まで


# ---------------------------------------------------------------------------
# ヘルパ
# ---------------------------------------------------------------------------
def safe_name(name):
    """パス区切りを除去した安全なファイル/フォルダ名にする。"""
    name = os.path.basename(name or "")
    return name.replace("\\", "").replace("/", "").strip()


def _long(path):
    """Windowsの260文字パス上限を回避する拡張パス（\\\\?\\）へ変換。他OSはそのまま。
       深い階層の長いPDF名で rmtree が失敗するのを防ぐ（transcribe_docs と同じ流儀）。"""
    if os.name != "nt":
        return path
    ap = os.path.abspath(path)
    if ap.startswith("\\\\?\\"):
        return ap
    if ap.startswith("\\\\"):
        return "\\\\?\\UNC\\" + ap[2:]
    return "\\\\?\\" + ap


def _safe_listdir(d):
    try:
        return os.listdir(d)
    except OSError:
        return []


def _doc_prefixes():
    """docs_config.json の各書類フォルダ名（2種関節系PRP / 3種筋腱靭帯系PRP / SOP）を出現順で返す。
       出力フォルダ名 '<書類種別>_<ヒアリングシート名>' を分解する接頭辞に使う。"""
    try:
        cfg = json.load(io.open(DOCS_CONFIG, encoding="utf-8"))
        names = [d.get("folder", "") for d in cfg.get("documents", {}).values() if d.get("folder")]
        if names:
            return names
    except Exception:
        pass
    return ["2種関節系PRP", "3種筋腱靭帯系PRP", "SOP"]


def _split_output_name(name, prefixes):
    """出力フォルダ名を (書類種別, ヒアリングシート名) に分解。該当接頭辞が無ければ (None, name)。"""
    for p in prefixes:
        if name.startswith(p + "_"):
            return p, name[len(p) + 1:]
    return None, name


def _short_type(kind):
    """表示用に書類種別を短縮（2種関節系PRP→2種 / SOP→SOP）。"""
    if not kind:
        return "その他"
    import re
    m = re.match(r"^(\d+種)", kind)
    if m:
        return m.group(1)
    if "SOP" in kind:
        return "SOP"
    return kind


def input_kind(name):
    """ファイル名から種別を判定（ヒアリングシート/略歴書/その他）。transcribe.py の選別と同じ流儀。"""
    b = os.path.basename(name or "")
    if "ヒアリング" in b:
        return "hearing"
    if "略歴" in b:
        return "rireki"
    return "other"


def list_inputs():
    files = [f for f in glob.glob(os.path.join(DIR_INPUT, "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    files.sort(key=os.path.getmtime, reverse=True)
    return [{
        "name": os.path.basename(f),
        "kind": input_kind(f),
        "mtime": datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M"),
        "size_kb": round(os.path.getsize(f) / 1024),
    } for f in files]


def detect_hearing():
    """01_input からヒアリングシート本体を自動判定（find_hearing と同じ優先順位）。"""
    cands = [f for f in glob.glob(os.path.join(DIR_INPUT, "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    named = [c for c in cands if "ヒアリング" in os.path.basename(c)]
    pool = named or [c for c in cands if "略歴" not in os.path.basename(c)] or cands
    pool.sort(key=os.path.getmtime, reverse=True)
    return os.path.basename(pool[0]) if pool else ""


def list_outputs():
    """02_output 直下のフォルダを『ヒアリングシート単位（＝1回の実行）』でまとめて新しい順に返す。
       2種/3種/SOP は同じヒアリングシート名を末尾に共有するため、1グループに束ねて日付単位で扱う。"""
    prefixes = _doc_prefixes()
    groups = {}
    for name in _safe_listdir(DIR_OUTPUT):
        full = os.path.join(DIR_OUTPUT, name)
        if not os.path.isdir(full):
            continue
        kind, base = _split_output_name(name, prefixes)
        nfiles = 0
        for _root, _dirs, fs in os.walk(full):
            nfiles += len([x for x in fs if not x.startswith("~$")])
        ts = os.path.getmtime(full)
        g = groups.get(base)
        if g is None:
            g = groups[base] = {"key": base, "types": [], "files": 0, "mtime_ts": ts}
        if kind and kind not in g["types"]:
            g["types"].append(kind)
        g["files"] += nfiles
        g["mtime_ts"] = max(g["mtime_ts"], ts)

    items = list(groups.values())
    items.sort(key=lambda x: x["mtime_ts"], reverse=True)
    result = []
    for g in items:
        types = sorted(g["types"], key=lambda t: prefixes.index(t) if t in prefixes else 99)
        dt = datetime.datetime.fromtimestamp(g["mtime_ts"])
        result.append({
            "key": g["key"],
            "name": g["key"],
            "types": [_short_type(t) for t in types],
            "files": g["files"],
            "date": dt.strftime("%Y-%m-%d"),
            "mtime": dt.strftime("%Y-%m-%d %H:%M"),
        })
    return result


def list_logs():
    files = glob.glob(os.path.join(DIR_LOGS, "転記ログ_*.xlsx"))
    files.sort(key=os.path.getmtime, reverse=True)
    return [{
        "name": os.path.basename(f),
        "mtime": datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M"),
    } for f in files]


def cleanup_old_logs(days=30):
    """DIR_LOGS の 転記ログ_*.xlsx のうち、更新日時が days 日より古いものを削除（起動時掃除）。
       返り値: 削除件数。エラーは握りつぶして起動を止めない。"""
    removed = 0
    try:
        cutoff = datetime.datetime.now() - datetime.timedelta(days=days)
        for f in glob.glob(os.path.join(DIR_LOGS, "転記ログ_*.xlsx")):
            try:
                if datetime.datetime.fromtimestamp(os.path.getmtime(f)) < cutoff:
                    os.remove(f)
                    removed += 1
            except OSError:
                pass
    except Exception:
        pass
    return removed


def parse_log(log_path):
    """転記ログ_*.xlsx を解析して 統計・要確認行・実行情報 を返す。"""
    wb = openpyxl.load_workbook(log_path, read_only=True, data_only=True)
    result = {"stats": {}, "checks": [], "info": {}, "total": 0}

    if "転記ログ" in wb.sheetnames:
        ws = wb["転記ログ"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            idx = {str(h): i for i, h in enumerate(header) if h is not None}

            def g(row, key):
                i = idx.get(key)
                return row[i] if (i is not None and i < len(row)) else None

            stats = {}
            for row in rows:
                if row is None or all(c is None for c in row):
                    continue
                result["total"] += 1
                status = str(g(row, "処理結果") or "")
                stats[status] = stats.get(status, 0) + 1
                if str(g(row, "確認") or "") == "要確認":
                    result["checks"].append({
                        "doc": str(g(row, "文書") or ""),
                        "sheet": str(g(row, "シート") or ""),
                        "cell": str(g(row, "セル") or ""),
                        "var": str(g(row, "変数") or ""),
                        "content": str(g(row, "転記内容") or ""),
                        "status": status,
                        "note": str(g(row, "備考") or ""),
                    })
            result["stats"] = stats

    if "実行情報" in wb.sheetnames:
        ws2 = wb["実行情報"]
        for row in ws2.iter_rows(values_only=True):
            if row and row[0]:
                result["info"][str(row[0])] = str(row[1]) if len(row) > 1 and row[1] is not None else ""
    wb.close()
    return result


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------
@app.get("/api/state")
def api_state():
    return jsonify({
        "inputs": list_inputs(),
        "outputs": list_outputs(),
        "logs": list_logs(),
        "root": BASE,
    })


@app.post("/api/upload")
def api_upload():
    """ヒアリングシート・略歴書などを一括アップロード（.xlsx 複数可）→ 01_input へ保存。"""
    files = request.files.getlist("file")
    if not files:
        one = request.files.get("file")
        files = [one] if one else []
    saved, skipped = [], []
    for f in files:
        if not f or not f.filename:
            continue
        name = safe_name(f.filename)
        if not name.lower().endswith(".xlsx"):
            skipped.append(name)
            continue
        f.save(os.path.join(DIR_INPUT, name))
        saved.append(name)
    if not saved:
        return jsonify({"error": "アップロードできる .xlsx がありませんでした（拡張子をご確認ください）",
                        "skipped": skipped}), 400
    return jsonify({"saved": saved, "skipped": skipped})


@app.post("/api/delete")
def api_delete():
    """01_input のファイルを1件削除（入力セットをその都度整えるため）。"""
    data = request.get_json(silent=True) or {}
    name = safe_name(data.get("name", "") or request.form.get("name", ""))
    if not name:
        return jsonify({"error": "ファイル名がありません"}), 400
    path = os.path.join(DIR_INPUT, name)
    if not os.path.exists(path):
        return jsonify({"error": "ファイルが見つかりません"}), 404
    try:
        os.remove(path)
    except Exception as e:
        return jsonify({"error": "削除に失敗: %r" % e}), 500
    return jsonify({"ok": True, "name": name})


@app.post("/api/delete-all")
def api_delete_all():
    """01_input の入力ファイル(.xlsx)を一括削除（次の案件に切り替える前の一掃用）。"""
    files = [f for f in glob.glob(os.path.join(DIR_INPUT, "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    deleted, failed = [], []
    for path in files:
        try:
            os.remove(path)
            deleted.append(os.path.basename(path))
        except Exception as e:
            failed.append({"name": os.path.basename(path), "error": "%r" % e})
    if failed:
        return jsonify({"error": "一部の削除に失敗しました",
                        "deleted": deleted, "failed": failed}), 500
    return jsonify({"ok": True, "deleted": deleted})


@app.get("/api/run")
def api_run():
    """ヒアリングシートで transcribe.py を実行し、進捗をSSEで流す。
       file 未指定なら 01_input のヒアリングシートを自動判定（略歴書は自動で取り込まれる）。"""
    fname = safe_name(request.args.get("file", "")) or detect_hearing()
    hearing_path = os.path.join(DIR_INPUT, fname) if fname else ""

    @stream_with_context
    def generate():
        def sse(event, data):
            return "event: %s\ndata: %s\n\n" % (event, json.dumps(data, ensure_ascii=False))

        if not fname or not os.path.exists(hearing_path):
            yield sse("error", {"message": "ヒアリングシートが見つかりません（先にアップロードしてください）"})
            return

        yield sse("log", {"line": "▶ 実行開始: %s" % fname})

        env = dict(os.environ)
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        env["PYTHONUNBUFFERED"] = "1"

        try:
            proc = subprocess.Popen(
                [sys.executable, TRANSCRIBE, hearing_path],
                cwd=BASE, env=env,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                bufsize=1, universal_newlines=True,
                encoding="utf-8", errors="replace",
            )
        except Exception as e:
            yield sse("error", {"message": "起動に失敗: %r" % e})
            return

        for line in iter(proc.stdout.readline, ""):
            line = line.rstrip("\n")
            if line != "":
                yield sse("log", {"line": line})
        proc.stdout.close()
        code = proc.wait()

        if code != 0:
            yield sse("error", {"message": "転記処理が異常終了しました（コード %s）" % code})
            return

        # 最新ログを解析して結果を返す
        logs = list_logs()
        result = {}
        log_name = ""
        if logs:
            log_name = logs[0]["name"]
            try:
                result = parse_log(os.path.join(DIR_LOGS, log_name))
            except Exception as e:
                yield sse("log", {"line": "（ログ解析でエラー: %r）" % e})

        yield sse("done", {
            "log": log_name,
            "result": result,
            "outputs": list_outputs(),
        })

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/api/log")
def api_log():
    """過去ログの解析結果を返す。"""
    name = safe_name(request.args.get("name", ""))
    path = os.path.join(DIR_LOGS, name)
    if not name or not os.path.exists(path):
        return jsonify({"error": "ログが見つかりません"}), 404
    return jsonify(parse_log(path))


@app.get("/api/download-log")
def api_download_log():
    name = safe_name(request.args.get("name", ""))
    path = os.path.join(DIR_LOGS, name)
    if not name or not os.path.exists(path):
        return jsonify({"error": "ログが見つかりません"}), 404
    return send_file(path, as_attachment=True, download_name=name)


def _output_members(group):
    """指定グループ（ヒアリングシート名）に属する 02_output 直下フォルダ名の一覧。"""
    prefixes = _doc_prefixes()
    members = []
    for name in _safe_listdir(DIR_OUTPUT):
        if not os.path.isdir(os.path.join(DIR_OUTPUT, name)):
            continue
        _kind, base = _split_output_name(name, prefixes)
        if base == group:
            members.append(name)
    return members


@app.get("/api/download")
def api_download():
    """02_output のうち、指定グループ（日付＝ヒアリングシート単位）の 2種/3種/SOP を1つのZIPにまとめて返す。
       group: グループキー。後方互換で folder（単一フォルダ名）も可。"""
    group = request.args.get("group", "")
    folder = safe_name(request.args.get("folder", ""))
    if group:
        targets = _output_members(group)
        zip_base = group
    elif folder and os.path.isdir(os.path.join(DIR_OUTPUT, folder)):
        targets = [folder]
        zip_base = folder
    else:
        targets = []
        zip_base = ""
    if not targets:
        return jsonify({"error": "フォルダが見つかりません"}), 404

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for folder_name in targets:
            # 拡張パス（\\?\）で走査・読み取りする。深い階層の長いPDF名で 260 文字上限を
            # 超えると os.stat が WinError 3 で失敗するため（論文PDF等で発生）。
            walk_root = _long(os.path.join(DIR_OUTPUT, folder_name))
            for root, _dirs, files in os.walk(walk_root):
                for fn in files:
                    if fn.startswith("~$"):
                        continue
                    fp = os.path.join(root, fn)
                    arc = os.path.join(folder_name, os.path.relpath(fp, walk_root))
                    zf.write(fp, arc)
    buf.seek(0)
    dl = (safe_name(zip_base) or "output") + ".zip"
    return send_file(buf, as_attachment=True, download_name=dl, mimetype="application/zip")


@app.post("/api/output/delete")
def api_output_delete():
    """02_output の1グループ（日付＝ヒアリングシート単位／2種・3種・SOP一式）を削除。"""
    data = request.get_json(silent=True) or {}
    group = data.get("group", "") or request.form.get("group", "")
    if not group:
        return jsonify({"error": "対象がありません"}), 400
    targets = _output_members(group)
    if not targets:
        return jsonify({"error": "フォルダが見つかりません"}), 404
    deleted, failed = [], []
    for name in targets:
        full = os.path.join(DIR_OUTPUT, name)
        try:
            shutil.rmtree(_long(full))
            deleted.append(name)
        except Exception as e:
            failed.append({"name": name, "error": "%r" % e})
    if failed:
        return jsonify({"error": "一部の削除に失敗しました",
                        "deleted": deleted, "failed": failed}), 500
    return jsonify({"ok": True, "deleted": deleted})


# ---------------------------------------------------------------------------
# WEB転記（e-再生医療フォームへ自動入力）
#   既存の 99_data/src/web_fill.py（対話型）を一切変更せず、サブプロセスとして
#   起動し、stdin にEnter/qを送り、stdout をSSEでライブ表示する薄いラッパー。
#   ・ブラウザは各自のPC上に開く（127.0.0.1 のダッシュボードから起動）
#   ・送信はしない（web_fill.py の仕様どおり、入力＝下書きまで）
# ---------------------------------------------------------------------------
WEB_LOCK = threading.Lock()
WEB = {"proc": None, "mode": None}       # 同時に1セッションのみ


def _web_mapping():
    try:
        with io.open(WEB_MAPPING, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _playwright_ready():
    """web_fill.py と同じ Python で playwright パッケージが入っているか。"""
    try:
        import importlib.util
        return importlib.util.find_spec("playwright") is not None
    except Exception:
        return False


def _web_output_folder():
    """web_mapping の output_folder_contains に一致する最新の 02_output フォルダ名。"""
    m = _web_mapping()
    want = m.get("output_folder_contains", "")
    best = None
    try:
        for name in os.listdir(DIR_OUTPUT):
            full = os.path.join(DIR_OUTPUT, name)
            if os.path.isdir(full) and (not want or want in name):
                ts = os.path.getmtime(full)
                if best is None or ts > best[0]:
                    best = (ts, name)
    except OSError:
        pass
    return best[1] if best else ""


@app.get("/api/web/status")
def api_web_status():
    proc = WEB["proc"]
    running = proc is not None and proc.poll() is None
    m = _web_mapping()
    return jsonify({
        "playwright": _playwright_ready(),
        "url": m.get("url", ""),
        "output_folder": _web_output_folder(),
        "running": running,
        "mode": WEB["mode"] if running else "",
    })


def _sse(event, data):
    return "event: %s\ndata: %s\n\n" % (event, json.dumps(data, ensure_ascii=False))


def _child_env():
    env = dict(os.environ)
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    env["PYTHONUNBUFFERED"] = "1"
    return env


@app.get("/api/web/start")
def api_web_start():
    """web_fill.py を起動し、ブラウザを開いて自動入力の対話を開始。進捗をSSEで流す。
       mode=fill: タブごと手動送りの自動入力 / mode=auto: 全タブ一括入力→一時保存まで自動 /
       mode=dump: フォーム項目の抽出（メンテ用）。いずれも送信はしない。"""
    mode = request.args.get("mode", "fill")
    flag = {"dump": ["--dump"], "auto": ["--auto"]}.get(mode, [])
    args = [sys.executable, WEB_FILL] + flag

    @stream_with_context
    def generate():
        err = None
        proc = None
        WEB_LOCK.acquire()
        try:
            if WEB["proc"] is not None and WEB["proc"].poll() is None:
                err = "すでにWEB転記セッションが実行中です。先に「終了」してください。"
            elif not os.path.exists(WEB_FILL):
                err = "web_fill.py が見つかりません。"
            else:
                try:
                    proc = subprocess.Popen(
                        args, cwd=BASE, env=_child_env(),
                        stdin=subprocess.PIPE,
                        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        bufsize=1, universal_newlines=True,
                        encoding="utf-8", errors="replace",
                    )
                    WEB["proc"] = proc
                    WEB["mode"] = mode
                except Exception as e:
                    err = "起動に失敗: %r" % e
        finally:
            WEB_LOCK.release()

        if err:
            yield _sse("error", {"message": err})
            return

        yield _sse("ready", {"mode": mode})
        try:
            for line in iter(proc.stdout.readline, ""):
                line = line.rstrip("\n")
                if line != "":
                    yield _sse("log", {"line": line})
        finally:
            try:
                proc.stdout.close()
            except Exception:
                pass
            code = proc.wait()
            with WEB_LOCK:
                if WEB["proc"] is proc:
                    WEB["proc"] = None
                    WEB["mode"] = None
            yield _sse("done", {"code": code})

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/api/web/send")
def api_web_send():
    """実行中セッションの stdin に指示を送る。cmd=fill→Enter（このページを入力/次へ） / cmd=quit→q（終了）。"""
    cmd = "fill"
    if request.is_json:
        cmd = (request.get_json(silent=True) or {}).get("cmd", "fill")
    else:
        cmd = request.form.get("cmd", "fill")

    proc = WEB["proc"]
    if proc is None or proc.poll() is not None:
        return jsonify({"error": "WEB転記セッションが実行されていません。"}), 400
    try:
        proc.stdin.write("q\n" if cmd == "quit" else "\n")
        proc.stdin.flush()
    except Exception as e:
        return jsonify({"error": "送信に失敗: %r" % e}), 500
    return jsonify({"ok": True})


@app.post("/api/web/stop")
def api_web_stop():
    """セッションを強制停止（ブラウザも閉じる）。まず q を送って猶予を与え、無理なら terminate。"""
    proc = WEB["proc"]
    if proc is None or proc.poll() is not None:
        return jsonify({"ok": True})
    try:
        proc.stdin.write("q\n")
        proc.stdin.flush()
    except Exception:
        pass
    try:
        proc.wait(timeout=3)
    except Exception:
        try:
            proc.terminate()
        except Exception:
            pass
    return jsonify({"ok": True})


@app.get("/api/web/setup")
def api_web_setup():
    """初回準備：playwright（pip）と Chromium ブラウザを導入。進捗をSSEで流す。"""
    @stream_with_context
    def generate():
        steps = [
            ("Playwright を導入しています…", [sys.executable, "-m", "pip", "install", "playwright"]),
            ("Chromium ブラウザを導入しています（数分かかることがあります）…",
             [sys.executable, "-m", "playwright", "install", "chromium"]),
        ]
        for title, cmd in steps:
            yield _sse("log", {"line": "▶ " + title})
            try:
                proc = subprocess.Popen(
                    cmd, cwd=BASE, env=_child_env(),
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    bufsize=1, universal_newlines=True,
                    encoding="utf-8", errors="replace",
                )
            except Exception as e:
                yield _sse("error", {"message": "導入コマンドの起動に失敗: %r" % e})
                return
            for line in iter(proc.stdout.readline, ""):
                line = line.rstrip("\n")
                if line != "":
                    yield _sse("log", {"line": line})
            proc.stdout.close()
            if proc.wait() != 0:
                yield _sse("error", {"message": "導入に失敗しました。ネットワーク環境をご確認ください。"})
                return
        yield _sse("done", {"playwright": _playwright_ready()})

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/")
def index():
    return Response(INDEX_HTML, mimetype="text/html; charset=utf-8")


# INDEX_HTML は index_html.py から読み込む（見通しのため分離）
from index_html import INDEX_HTML  # noqa: E402


if __name__ == "__main__":
    import webbrowser
    import threading

    port = int(os.environ.get("PRP_DASH_PORT", "8765"))
    url = "http://127.0.0.1:%d" % port
    print("=" * 60)
    print(" PRP 自動転記ツール ダッシュボード")
    print(" ブラウザで開く: %s" % url)
    print(" 終了するには、この画面で Ctrl + C")
    print("=" * 60)

    # 起動時のログ掃除：30日より古い転記ログを自動削除
    _removed = cleanup_old_logs(30)
    if _removed:
        print(" ※ 30日より古いログ %d 件を自動削除しました。" % _removed)

    # 起動直後にブラウザを開く（reloader無効時のみ）
    if not os.environ.get("WERKZEUG_RUN_MAIN"):
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    app.run(host="127.0.0.1", port=port, debug=False, threaded=True)
