# -*- coding: utf-8 -*-
"""
確定実行（Step2）: 下書き(経費入力表_取込*.xlsx)の『実行=☑』行を確定する。
  ① 原本PDFを  YYYYMMDD_相手先_品目_金額円.pdf にリネームし、年月フォルダへ配置
  ② 経費入力表.xlsx の月シート「経費(N)」へ 日付/相手先/内容/金額 を追記
安全策:
  ・マスターは書込前に必ずタイムスタンプ付きでバックアップ
  ・確定した下書き行は『済』に更新（二重計上を防止）
  ・⚠要手入力／実行未チェックの行はスキップ
※ 送信・課金なし。ローカルのファイル操作のみ。
"""
import os
import io
import re
import glob
import shutil
import datetime

import openpyxl

import pipeline as P

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DIR_OUT = os.path.join(ROOT, "02_output")
SHEET = "領収書取込"
# 下書きの列（receipt_ocr.py と一致）
C_MON, C_DATE, C_AITE, C_NAIYO, C_PAY = 1, 2, 6, 7, 9
C_STATE, C_NAME, C_YM, C_DO, C_SRC = 12, 17, 18, 19, 20
MASTER_HEADERS = ["月", "日付", "GWS", "立替", "清算", "相手先", "内容",
                  "収入", "支払", "差引（残額）"]


def _backup(path):
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = "%s.backup_%s%s" % (os.path.splitext(path)[0], stamp, os.path.splitext(path)[1])
    shutil.copy2(path, bak)
    return bak


def _first_empty_row(ws):
    r = 5
    while ws.cell(r, 1).value not in (None, "") or ws.cell(r, 6).value not in (None, ""):
        r += 1
    return r


def _ensure_master_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(MASTER_HEADERS, 1):
        ws.cell(4, c, h)
    return ws


def append_to_master(wb, sheet_fmt, dt, month, aite, naiyo, amount):
    ws = _ensure_master_sheet(wb, P.month_sheet_name(dt, sheet_fmt))
    r = _first_empty_row(ws)
    ws.cell(r, 1, month)
    b = ws.cell(r, 2, dt)
    b.number_format = "yyyy/m/d"
    ws.cell(r, 6, aite)
    ws.cell(r, 7, naiyo)
    ws.cell(r, 9, int(amount))
    ws.cell(r, 10, "=H%d-I%d" % (r, r))
    return ws.title, r


def place_file(src, plan_name, ym, output_root, action):
    """原本を 年月フォルダへ リネームして配置。実施したパスを返す（不可ならNone）。"""
    if not (src and plan_name and ym) or not os.path.exists(src):
        return None
    folder = os.path.join(output_root, ym)
    os.makedirs(folder, exist_ok=True)
    dest = P.unique_path(folder, plan_name)
    if action == "move":
        shutil.move(src, dest)
    else:
        shutil.copy2(src, dest)
    return dest


def process_draft(path, cfg, wb_master, logs):
    wb = openpyxl.load_workbook(path)
    if SHEET not in wb.sheetnames:
        return 0, 0
    ws = wb[SHEET]
    done_x = done_f = 0
    for r in range(3, ws.max_row + 1):
        do = str(ws.cell(r, C_DO).value or "").strip()
        if do not in ("☑", "1", "○", "レ", "V", "v", "x", "X"):
            continue                                   # 未チェック/『済』はスキップ
        dt = P.parse_date(ws.cell(r, C_DATE).value)
        amount = re.sub(r"[^0-9]", "", str(ws.cell(r, C_PAY).value or ""))
        if not dt or not amount:
            logs.append("  ・skip(r%d): 日付か金額が空" % r)
            continue
        aite = ws.cell(r, C_AITE).value or ""
        naiyo = ws.cell(r, C_NAIYO).value or ""
        # ① 経費入力表へ蓄積
        sname, mr = append_to_master(wb_master, cfg["month_sheet_fmt"], dt, dt.month,
                                     aite, naiyo, amount)
        done_x += 1
        # ② 原本のリネーム＆年月配置（提案名がある＝1PDF1明細のときのみ）
        dest = place_file(ws.cell(r, C_SRC).value, ws.cell(r, C_NAME).value,
                          ws.cell(r, C_YM).value, cfg["output_root"], cfg["file_action"])
        if dest:
            done_f += 1
            logs.append("  ○ %s → %s（%s r%d）" %
                        (os.path.basename(str(ws.cell(r, C_SRC).value)),
                         os.path.basename(dest), sname, mr))
        else:
            logs.append("  ○ 蓄積のみ: %s ｜ %s円（%s r%d）" %
                        (str(aite)[:16], amount, sname, mr))
        ws.cell(r, C_DO, "済")                          # 二重計上防止
    wb.save(path)                                       # 『済』を保存
    return done_x, done_f


def main():
    cfg = P.load_settings()
    print("=== 確定実行（リネーム・年月配置・経費入力表へ蓄積）===")
    print("  蓄積先: %s" % cfg["master_xlsx"])
    print("  保存先: %s" % cfg["output_root"])
    print("  原本の扱い: %s" % ("移動" if cfg["file_action"] == "move" else "コピー"))
    drafts = sorted(glob.glob(os.path.join(DIR_OUT, "経費入力表_取込*.xlsx")))
    if not drafts:
        print("下書き(経費入力表_取込*.xlsx)が 02_output にありません。先に読取実行してください。")
        return
    if not os.path.exists(cfg["master_xlsx"]):
        # 蓄積先が無ければ、元フォーマット(テンプレ)からコピーして作る（保存用＋経費(N)を維持）。
        os.makedirs(os.path.dirname(cfg["master_xlsx"]), exist_ok=True)
        tmpl = cfg.get("template_xlsx", "")
        if tmpl and os.path.exists(tmpl):
            shutil.copy2(tmpl, cfg["master_xlsx"])
            print("  ※ 蓄積先が無いのでテンプレから作成: %s" % cfg["master_xlsx"])
            wb_master = openpyxl.load_workbook(cfg["master_xlsx"])
        else:
            wb_master = openpyxl.Workbook()
            wb_master.remove(wb_master.active)
            print("  ※ 蓄積先もテンプレも無いため空で新規作成します。")
        bak = None
    else:
        bak = _backup(cfg["master_xlsx"])
        print("  バックアップ作成: %s" % bak)
        wb_master = openpyxl.load_workbook(cfg["master_xlsx"])
    logs = []
    tot_x = tot_f = 0
    for d in drafts:
        print("\n▼ 下書き: %s" % os.path.basename(d))
        x, f = process_draft(d, cfg, wb_master, logs)
        tot_x += x
        tot_f += f
    os.makedirs(os.path.dirname(cfg["master_xlsx"]), exist_ok=True)
    wb_master.save(cfg["master_xlsx"])
    out = ["=== 確定実行レポート (%s) ===" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M")]
    out += logs
    out.append("")
    out.append("→ 経費入力表へ %d件 蓄積 ／ 原本 %d件 を年月フォルダへ配置。" % (tot_x, tot_f))
    if bak:
        out.append("→ マスターのバックアップ: %s" % bak)
    out.append("→ 蓄積先: %s" % cfg["master_xlsx"])
    text = "\n".join(out)
    print("\n" + text)
    with io.open(os.path.join(DIR_OUT, "確定実行レポート.txt"), "w", encoding="utf-8") as fh:
        fh.write(text)


if __name__ == "__main__":
    main()
