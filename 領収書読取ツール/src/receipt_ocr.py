# -*- coding: utf-8 -*-
"""
領収書読取ツール（ローカル・API課金なし）
  スキャンPDF(複数レシート/回転あり) → 個別に分割 → 向き補正 → OCR(Tesseract日本語)
  → {日付・相手先・内容・金額・種別} を抽出 → 経費入力表フォーマットのExcelへ「追記」。
  手書き/低信頼は「⚠要手入力」でフラグし黄色で色付け。新規追加行は分かるように印を付ける。

  ※ Tesseractは同梱ポータブル(_ocr/tesseract)を自動使用（管理者不要）。
  ※ 送信・外部送信は一切しない。全てローカル処理。
"""
import os
import io
import re
import glob
import datetime
import subprocess

import fitz                      # PyMuPDF
import numpy as np
import cv2
from PIL import Image
import pytesseract
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ---- パス設定 -------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DIR_IN = os.path.join(ROOT, "01_input")
DIR_OUT = os.path.join(ROOT, "02_output")
DIR_WORK = os.path.join(ROOT, "03_work")
TESS_DIR = os.path.join(ROOT, "_ocr", "tesseract")
TESS_EXE = os.path.join(TESS_DIR, "tesseract.exe")
TESSDATA = os.path.join(TESS_DIR, "tessdata")
BASE_XLSX = os.path.join(ROOT, "経費入力表.xlsx")          # 育てていくベース（無ければ新規作成）
OUT_XLSX = os.path.join(DIR_OUT, "経費入力表_取込.xlsx")   # 実際に追記していく出力
SHEET = "領収書取込"

DPI = 300
LOWCONF = 55                    # OCR平均信頼度がこれ未満なら「要確認/手書きの可能性」

for d in (DIR_IN, DIR_OUT, DIR_WORK):
    os.makedirs(d, exist_ok=True)

if os.path.exists(TESS_EXE):
    pytesseract.pytesseract.tesseract_cmd = TESS_EXE
    os.environ["TESSDATA_PREFIX"] = TESSDATA

# ローカルLLM（Ollama）を使えるなら主軸に。使えなければOCR/ルールのみで動く。
LLM_ON = False
L = None
if os.environ.get("RECEIPT_USE_LLM", "1") != "0":
    try:
        import llm_extract as L
    except Exception:
        L = None


def _ensure_llm(wait_sec=25):
    """Ollamaの起動完了を最大 wait_sec 秒待つ（起動直後でも取りこぼさないため）。"""
    global LLM_ON
    if L is None:
        LLM_ON = False
        return
    import time
    ok, _ = L.ready()
    waited = 0
    while not ok and waited < wait_sec:
        if waited == 0:
            print("  Ollama(画像LLM)の起動を待っています…（最大%d秒）" % wait_sec)
        time.sleep(2)
        waited += 2
        ok, _ = L.ready()
    LLM_ON = ok


def _imwrite_u(path, img):
    """日本語パスでも確実に保存（cv2.imwriteはWindowsの非ASCIIパスで失敗するため）。"""
    ext = os.path.splitext(path)[1] or ".png"
    ok, buf = cv2.imencode(ext, img)
    if ok:
        buf.tofile(path)
    return ok


def _to_int_amount(v):
    if isinstance(v, (int, float)):
        return int(v)
    s = re.sub(r"[^\d]", "", str(v or ""))
    return int(s) if s else ""


def _merge_fields(llm, rule):
    """LLM(主) と ルール/OCR(補) を統合。戻り値: (fields, 種別)。"""
    f = {"日付": rule.get("日付", ""), "相手先": rule.get("相手先", ""),
         "内容": "", "金額": rule.get("金額", "")}
    kind = rule.get("内容", "")                       # ルールの「内容」は種別相当(_kind)
    if llm and not llm.get("_error"):
        if isinstance(llm.get("日付"), str) and llm["日付"].strip():
            f["日付"] = llm["日付"].strip()
        if isinstance(llm.get("相手先"), str) and llm["相手先"].strip():
            f["相手先"] = llm["相手先"].strip()
        amt = _to_int_amount(llm.get("金額"))
        if amt not in ("", 0):
            f["金額"] = amt
        c = llm.get("内容")
        if isinstance(c, str) and c.strip() and not ("但し書き" in c and "要約" in c):
            f["内容"] = c.strip()
        k = llm.get("種別")
        if isinstance(k, str) and k.strip() and "から1つ" not in k:
            kind = k.strip()
    if not f["内容"]:
        f["内容"] = kind
    return f, kind


# ---- 前処理・分割 ---------------------------------------------------------
def render_page(pdf_path, page_index, dpi=DPI):
    doc = fitz.open(pdf_path)
    pg = doc[page_index]
    pix = pg.get_pixmap(matrix=fitz.Matrix(dpi / 72.0, dpi / 72.0))
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR) if pix.n == 4 else \
        (cv2.cvtColor(img, cv2.COLOR_RGB2BGR) if pix.n == 3 else cv2.cvtColor(img, cv2.COLOR_GRAY2BGR))
    n = doc.page_count
    doc.close()
    return img, n


def _merge_boxes(boxes, w, h, gap_x=0.02, gap_y=0.02):
    gx, gy = w * gap_x, h * gap_y
    changed = True
    while changed:
        changed = False
        out, used = [], [False] * len(boxes)
        for i in range(len(boxes)):
            if used[i]:
                continue
            a = boxes[i][:]
            for j in range(i + 1, len(boxes)):
                if used[j]:
                    continue
                b = boxes[j]
                if (a[0] - gx < b[2] and b[0] - gx < a[2] and a[1] - gy < b[3] and b[1] - gy < a[3]):
                    a = [min(a[0], b[0]), min(a[1], b[1]), max(a[2], b[2]), max(a[3], b[3])]
                    used[j] = True
                    changed = True
            used[i] = True
            out.append(a)
        boxes = out
    return boxes


def detect_receipts(img):
    """1ページ内の各レシートの矩形を返す（余白パディング込み）。"""
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    th = cv2.medianBlur(th, 3)
    k = cv2.getStructuringElement(cv2.MORPH_RECT, (max(15, w // 60), max(15, h // 80)))
    closed = cv2.morphologyEx(th, cv2.MORPH_CLOSE, k, iterations=2)
    cnts, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    boxes = []
    for c in cnts:
        x, y, ww, hh = cv2.boundingRect(c)
        if ww * hh < w * h * 0.02 or ww < w * 0.08 or hh < h * 0.05:
            continue
        boxes.append([x, y, x + ww, y + hh])
    boxes = _merge_boxes(boxes, w, h)
    # パディング＋ページ順（上→下、左→右）
    pad = int(min(w, h) * 0.012)
    out = []
    for b in boxes:
        out.append([max(0, b[0] - pad), max(0, b[1] - pad),
                    min(w, b[2] + pad), min(h, b[3] + pad)])
    out = sorted(out, key=lambda b: (round(b[1] / (h * 0.15)), b[0]))
    # 分割できない（1枚もの）ときはページ全体を1件に
    if not out:
        out = [[0, 0, w, h]]
    return out


def auto_rotate(crop):
    """Tesseract OSD で 0/90/180/270 の向きを判定して補正した画像を返す。"""
    try:
        osd = pytesseract.image_to_osd(crop, config="--psm 0")
        m = re.search(r"Rotate:\s*(\d+)", osd)
        deg = int(m.group(1)) if m else 0
    except Exception:
        deg = 0
    if deg == 90:
        return cv2.rotate(crop, cv2.ROTATE_90_CLOCKWISE)
    if deg == 180:
        return cv2.rotate(crop, cv2.ROTATE_180)
    if deg == 270:
        return cv2.rotate(crop, cv2.ROTATE_90_COUNTERCLOCKWISE)
    return crop


# ---- OCR・手書き判定 ------------------------------------------------------
def ocr_crop(crop):
    """(text, mean_conf) を返す。"""
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 5, 40, 40)
    pil = Image.fromarray(gray)
    text = pytesseract.image_to_string(pil, lang="jpn", config="--psm 6")
    try:
        data = pytesseract.image_to_data(pil, lang="jpn", config="--psm 6",
                                         output_type=pytesseract.Output.DICT)
        confs = [int(c) for c, t in zip(data["conf"], data["text"]) if str(c).lstrip("-").isdigit()
                 and int(c) >= 0 and t.strip()]
        mean_conf = sum(confs) / len(confs) if confs else 0
    except Exception:
        mean_conf = 0
    return text, mean_conf


def red_ink_ratio(crop):
    """赤ペン（手書き注記）の割合。多ければ手書き書き込みの可能性。"""
    hsv = cv2.cvtColor(crop, cv2.COLOR_BGR2HSV)
    m1 = cv2.inRange(hsv, (0, 80, 80), (10, 255, 255))
    m2 = cv2.inRange(hsv, (170, 80, 80), (180, 255, 255))
    red = cv2.countNonZero(m1 | m2)
    return red / (crop.shape[0] * crop.shape[1])


def judge_handwriting(text, mean_conf, crop, llm=None, fields=None):
    """(is_flag, reason)。要手入力なら True。
       LLMが使える時はLLMの手書き判定を優先（印刷レシートの赤印で誤検知しない）。
       金額が読めない場合は種類を問わず「要確認」。"""
    reasons = []
    llm_ok = bool(llm) and not llm.get("_error")
    if llm_ok and llm.get("手書き") is True:
        reasons.append("手書きが主体（LLM判定）")
    # ★安全網：OCRがほとんど読めない＝手書き/低画質の可能性。
    #   LLMが自信満々に“幻覚”で読む危険があるため、LLM有無に関わらず必ず要確認にする。
    if mean_conf < 45:
        reasons.append("OCRが読めない(信頼度%.0f)＝手書き/低画質の可能性" % mean_conf)
    if len(re.sub(r"\s", "", text)) < 10:
        reasons.append("機械文字がほとんど無い")
    if fields is not None and (fields.get("金額") in ("", 0, None)):
        reasons.append("金額が読み取れない")
    return (len(reasons) > 0), "／".join(reasons)


# ---- 項目抽出（v1：ルールベース。後でOllamaで強化予定） -------------------
_BRANDS = ["セブン-イレブン", "セブンイレブン", "ファミリーマート", "FamilyMart", "ローソン",
           "厚生労働省年金局", "日本年金機構", "法務局", "民事法務協会", "印刷産業連合会",
           "Amazon", "アマゾン", "楽天"]


def _to_iso_date(text):
    m = re.search(r"(20\d{2})[/\-年\.](\d{1,2})[/\-月\.](\d{1,2})", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return "%04d-%02d-%02d" % (y, mo, d)
        except Exception:
            return ""
    m = re.search(r"令和\s*(\d{1,2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if m:
        y = 2018 + int(m.group(1))
        return "%04d-%02d-%02d" % (y, int(m.group(2)), int(m.group(3)))
    return ""


def _amount(text):
    """合計らしき金額を推定。¥付き・「合計/計」近傍を優先。箱文字(¥ 2 6 7 7 7)にも対応。
       電話番号・年月日・各種番号は金額として拾わない。"""
    cands = []   # (金額, 優先度)

    def push(digits, pri):
        digits = re.sub(r"[^\d]", "", digits)
        if 2 <= len(digits) <= 8:
            cands.append((int(digits), pri))

    # ① ¥ の直後（スペース区切りの箱文字も許容）
    for m in re.finditer(r"[¥￥]\s*([0-9][0-9,\s]{0,14})", text):
        push(m.group(1), 3)
    # ② 「合計/総計/お会計/計」の近傍
    for m in re.finditer(r"(合\s*計|総\s*計|お会計|計)\D{0,8}([0-9][0-9,\s]{0,12})", text):
        push(m.group(2), 2)
    # ③ フォールバック：電話/番号/年月日を含む行は除外して数字を拾う
    if not cands:
        for line in text.splitlines():
            if re.search(r"(電話|TEL|No|番号|連番|管理|受付|年|月|日|コード)", line):
                continue
            for m in re.finditer(r"([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{3,6})", line):
                push(m.group(1), 1)
    if not cands:
        return ""
    best = max(cands, key=lambda c: (c[1], c[0]))   # 優先度→金額
    return best[0]


def _vendor(text):
    for b in _BRANDS:
        if b in text:
            return b
    for line in [l.strip() for l in text.splitlines() if l.strip()]:
        if any(x in line for x in ("店", "株式会社", "協会", "機構", "局")):
            return line[:24]
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    return lines[0][:24] if lines else ""


def _kind(text):
    table = [("印紙", "印紙代"), ("証明書", "証明書代"), ("年金", "社会保険料"),
             ("保険", "社会保険料"), ("交通", "交通費"), ("駐車", "駐車場代"),
             ("切手", "郵送費"), ("通信", "通信費")]
    for kw, name in table:
        if kw in text:
            return name
    return ""


def extract_fields(text):
    return {"日付": _to_iso_date(text), "相手先": _vendor(text),
            "内容": _kind(text), "金額": _amount(text)}


# ---- Excel 出力（経費入力表と同じ列に合わせて“そのままコピペ”できる体裁） ----
#   A〜J = 経費入力表と同じ列（貼り付け用・無色）／ K〜P = 補助（確認・状態・画像・OCR）
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")     # 新規(補助列)=薄い黄
FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")    # 要手入力(補助列)=赤
OK_FILL = PatternFill("solid", fgColor="E2EFDA")      # 状態OK=薄緑
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
COPY_HEAD_FILL = PatternFill("solid", fgColor="FCE4D6")  # 貼付対象(A〜J)の見出し=薄橙
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
# 経費入力表の列並びに一致（A〜J）＋補助（K〜P）
COLS = ["月", "日付", "GWS", "立替", "清算", "相手先", "内容", "収入", "支払", "差引（残額）",
        "確認", "状態", "元PDF", "No", "画像", "OCR抜粋"]
COPY_COLS = 10                                        # A〜J が経費入力表と同じ＝貼り付け対象
IMG_COL = 15                                          # O列に画像
WRAP_COLS = {7, 16}                                   # 内容 / OCR抜粋 は折り返し


def _open_out(out_xlsx):
    if os.path.exists(out_xlsx):
        return openpyxl.load_workbook(out_xlsx)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def _ensure_sheet(wb):
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        if ws.cell(2, 1).value == "月" and ws.cell(2, 6).value == "相手先":
            return ws                                  # 既に新レイアウト
        # 旧レイアウト → 退避（データは残す）して新規作成
        old = SHEET + "_旧"
        i = 2
        while old in wb.sheetnames:
            old = "%s_旧%d" % (SHEET, i)
            i += 1
        ws.title = old
    ws = wb.create_sheet(SHEET, 0)
    # 1行目：使い方メモ（貼り付け範囲の案内）
    ws.cell(1, 1, "▼この行の下がデータ。A〜I列（月〜支払）を選んでコピー→経費入力表の空き行に貼り付け。"
                  "／K〜P列は確認用（画像・OCR）。赤=要手入力。")
    ws.cell(1, 1).font = Font(bold=True, color="C00000")
    # 2行目：見出し
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(2, j, c)
        cell.fill = COPY_HEAD_FILL if j <= COPY_COLS else HEAD_FILL
        cell.font = Font(bold=True)
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center")
    widths = [5, 12, 6, 6, 6, 22, 28, 8, 10, 11,   # A〜J
              6, 15, 24, 5, 45, 40]                # K〜P（O=画像列を広めに）
    for j, wd in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = wd
    ws.freeze_panes = "A3"
    return ws


def append_rows(entries, out_xlsx):
    wb = _open_out(out_xlsx)
    ws = _ensure_sheet(wb)
    start = ws.max_row + 1
    batch = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for e in entries:
        r = ws.max_row + 1
        f = e["fields"]
        mon = (f["日付"][5:7].lstrip("0") if f.get("日付") else "")
        state = "⚠要手入力" if e["flag"] else "OK"
        # A〜J（経費入力表と同じ並び）／ K〜P（補助）
        vals = [mon, f.get("日付", ""), "", "", "", f.get("相手先", ""),
                f.get("内容", ""), "", f.get("金額", ""), "",
                "☐", state, e["pdf"], e["no"], "", e["ocr_excerpt"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=(j in WRAP_COLS))
        ws.cell(r, 11).alignment = Alignment(horizontal="center", vertical="top")  # 確認☐
        # 色付けは補助列(K〜P)だけ＝A〜Jは無色のまま貼り付けられる
        helper_fill = FLAG_FILL if e["flag"] else NEW_FILL
        for j in range(11, len(COLS) + 1):
            ws.cell(r, j).fill = helper_fill
        ws.cell(r, 12).fill = FLAG_FILL if e["flag"] else OK_FILL                  # 状態セル
        # 切り出し画像（O列）：大きめ＋アスペクト比維持で読みやすく。
        # 行高を画像に合わせ、縦長でも下の行にはみ出さない（重なり防止）。
        if e.get("thumb") and os.path.exists(e["thumb"]):
            try:
                im = XLImage(e["thumb"])
                ratio = e.get("thumb_ratio") or 1.4          # 高さ/幅
                box_w, box_h = 300, 460                       # 表示上限(px)
                w = box_w
                h = int(w * ratio)
                if h > box_h:                                # 縦長すぎるものは高さ基準で縮小
                    h = box_h
                    w = max(60, int(h / ratio))
                im.width, im.height = w, h
                ws.row_dimensions[r].height = h * 0.75 + 8    # 画像高さ(px)→行高(pt)＋余白
                ws.add_image(im, "%s%d" % (openpyxl.utils.get_column_letter(IMG_COL), r))
            except Exception:
                pass
    wb.save(out_xlsx)
    return start, ws.max_row, batch


# ---- メイン ---------------------------------------------------------------
def check_ready():
    if not os.path.exists(TESS_EXE):
        return False, "Tesseract(_ocr/tesseract)が見つかりません。セットアップを実行してください。"
    if not os.path.exists(os.path.join(TESSDATA, "jpn.traineddata")):
        return False, "日本語データ(jpn.traineddata)がありません。セットアップを実行してください。"
    return True, "OK"


def process_pdf(pdf_path, report):
    base = os.path.basename(pdf_path)
    img, npages = render_page(pdf_path, 0)
    entries = []
    for pi in range(npages):
        if pi > 0:
            img, _ = render_page(pdf_path, pi)
        boxes = detect_receipts(img)
        for idx, b in enumerate(boxes, 1):
            crop = auto_rotate(img[b[1]:b[3], b[0]:b[2]])
            no = "%d-%d" % (pi + 1, idx) if npages > 1 else str(idx)
            thumb = os.path.join(DIR_WORK, "%s_p%d_%d.png" % (base.replace(".pdf", ""), pi + 1, idx))
            _imwrite_u(thumb, crop)                        # 日本語パス対応で保存（サムネイル用）
            text, conf = ocr_crop(crop)                    # OCR＝手書き判定の補助＋フォールバック
            rule = extract_fields(text)
            llm = L.extract(thumb) if LLM_ON else {}       # LLM＝主軸（画像を見て抽出）
            fields, kind = _merge_fields(llm, rule)
            flag, reason = judge_handwriting(text, conf, crop, llm, fields)
            ratio = crop.shape[0] / max(1, crop.shape[1])
            excerpt = " ".join(text.split())[:120]
            entries.append({"pdf": base, "no": no, "fields": fields, "種別": kind, "flag": flag,
                            "reason": reason, "conf": conf, "thumb": thumb,
                            "thumb_ratio": ratio, "ocr_excerpt": excerpt})
            tag = ("⚠要手入力(%s)" % reason) if flag else "OK"
            report.append("  %s No.%s: %s ｜ %s ｜ %s ｜ %s円 ｜ %s" %
                          (base, no, tag, fields["日付"], fields["相手先"][:18],
                           fields["金額"], fields["内容"][:20]))
    return entries


def _safe_name(s):
    """フォルダ名をファイル名に使えるよう禁止文字を置換。"""
    return re.sub(r'[\\/:*?"<>|]', '_', (s or "")).strip() or "folder"


def run_group(label, pdfs, out_xlsx, report_name):
    """1グループ（直下PDF or 1フォルダ）を処理して、そのグループ専用のExcel/レポートに出力。"""
    report = ["=== 取込レポート (%s) ===" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M")]
    if label:
        report.append("対象フォルダ: %s" % label)
    all_entries = []
    for p in pdfs:
        report.append("▼ %s" % os.path.basename(p))
        all_entries += process_pdf(p, report)
    s, e, batch = append_rows(all_entries, out_xlsx)
    n_flag = sum(1 for x in all_entries if x["flag"])
    report.append("")
    report.append("→ 今回 %d件 追加（%s）。うち ⚠要手入力=%d件。" % (len(all_entries), batch, n_flag))
    report.append("→ 出力: %s（シート「%s」・黄=新規/赤=要手入力）" % (out_xlsx, SHEET))
    if n_flag:
        report.append("")
        report.append("⚠ 手書き/低信頼を検知した項目（内容をご確認のうえ手入力してください）:")
        for x in all_entries:
            if x["flag"]:
                report.append("   ・%s No.%s（%s）" % (x["pdf"], x["no"], x["reason"]))
    out = "\n".join(report)
    print(out)
    with io.open(os.path.join(DIR_OUT, report_name), "w", encoding="utf-8") as f:
        f.write(out)
    return len(all_entries), n_flag


def main():
    ok, msg = check_ready()
    print("=== 領収書読取ツール ===")
    if not ok:
        print("★ " + msg)
        return
    _ensure_llm()                                      # Ollamaの起動完了を待ってから判定
    print("  抽出エンジン: %s" % ("ローカルLLM(%s)＋OCR" % L.DEFAULT_MODEL if LLM_ON
                                   else "OCR/ルールのみ（Ollama未起動 → 精度は控えめ）"))
    # 入力の受け付け方:
    #   ・01_input 直下のPDF        → 1つのExcel（経費入力表_取込.xlsx）にまとめる
    #   ・01_input 直下のサブフォルダ → フォルダごとに別Excel（経費入力表_取込_<フォルダ名>.xlsx）
    top_pdfs = sorted(glob.glob(os.path.join(DIR_IN, "*.pdf")))
    subdirs = sorted(d for d in glob.glob(os.path.join(DIR_IN, "*")) if os.path.isdir(d))
    groups = []                                        # (表示名, pdf群, 出力xlsx, レポート名)
    if top_pdfs:
        groups.append(("", top_pdfs, OUT_XLSX, "取込レポート.txt"))
    for d in subdirs:
        name = os.path.basename(d)
        pdfs = sorted(glob.glob(os.path.join(d, "**", "*.pdf"), recursive=True))
        if not pdfs:
            continue
        safe = _safe_name(name)
        groups.append((name, pdfs,
                       os.path.join(DIR_OUT, "経費入力表_取込_%s.xlsx" % safe),
                       "取込レポート_%s.txt" % safe))
    if not groups:
        print("01_input にPDF（またはPDF入りのフォルダ）を入れてください。")
        return
    total = total_flag = 0
    for label, pdfs, out_xlsx, rep in groups:
        head = ("フォルダ「%s」" % label) if label else "直下のPDF"
        print("\n■ %s（%d件）→ %s" % (head, len(pdfs), os.path.basename(out_xlsx)))
        n, nf = run_group(label, pdfs, out_xlsx, rep)
        total += n
        total_flag += nf
    if len(groups) > 1:
        print("\n=== 全%dグループ 合計: %d件（うち ⚠要手入力 %d件）===" % (len(groups), total, total_flag))


if __name__ == "__main__":
    main()
